import secrets
import io
from datetime import datetime, timedelta
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, current_app, abort, jsonify, send_file)
from flask_login import current_user, login_required
from sqlalchemy import or_, false as sql_false, func
from models import (db, User, Application, Position, University, UniversityDepartment, UniversityMember,
                    ApplicationHistory, CompanyMember,
                    ROLE_STUDENT, ROLE_UNIVERSITY_COORD, ROLE_ADMIN,
                    ALL_STATUSES, STATUS_UNIV_PENDING, STATUS_NEW, STATUS_REVIEW, STATUS_INTERVIEW,
                    STATUS_OFFER, STATUS_HIRED, STATUS_REJECTED, SOURCES)
from helpers import university_coordinator_required, log_audit, send_email, push_notification, save_company_image

university_bp = Blueprint('university', __name__)


def _my_membership():
    if current_user.role == ROLE_ADMIN:
        return None
    q = UniversityMember.query.filter_by(user_id=current_user.id)
    if current_user.university_id:
        q = q.filter_by(university_id=current_user.university_id)
    return q.first()


def _my_university():
    """Return the University this coordinator belongs to, or None."""
    member = _my_membership()
    return member.university if member else None


@university_bp.route('/')
@university_coordinator_required
def dashboard():
    univ = _my_university()
    if not univ and current_user.role != ROLE_ADMIN:
        flash('You are not linked to any university yet.', 'warning')
        return render_template('university/dashboard.html', univ=None,
                               total_students=0, internship_apps=0,
                               active_apps=0, hired_count=0, recent_apps=[])

    membership = _my_membership()
    student_ids = _student_ids(univ, membership)
    total_students  = len(student_ids)
    internship_q    = _internship_apps(student_ids)
    internship_apps = internship_q.count()
    active_apps     = internship_q.filter(
        Application.status.in_([STATUS_UNIV_PENDING, STATUS_NEW, STATUS_REVIEW, STATUS_INTERVIEW, STATUS_OFFER])
    ).count()
    hired_count = internship_q.filter_by(status=STATUS_HIRED).count()
    recent_apps = (internship_q
                   .order_by(Application.applied_at.desc())
                   .limit(8).all())

    # Status breakdown
    status_counts = {s: internship_q.filter_by(status=s).count() for s in ALL_STATUSES}

    # ── Cohort intelligence ──────────────────────────────────────────────
    cohort = _cohort_insights(univ, membership, student_ids, internship_q)

    return render_template('university/dashboard.html',
        univ=univ,
        total_students=total_students,
        internship_apps=internship_apps,
        active_apps=active_apps,
        hired_count=hired_count,
        recent_apps=recent_apps,
        status_counts=status_counts,
        cohort=cohort,
        ALL_STATUSES=ALL_STATUSES)


@university_bp.route('/me', methods=['GET', 'POST'])
@login_required
def my_university():
    """My University — single page accessible to both coordinators and students.

    - Students: read-only view of their linked university.
    - Coordinators: same view plus an inline edit form, but only while the
      university has not yet been verified by an admin. Once verified, the
      form becomes read-only and a notice tells them to contact admins.
    - Admins/others: redirected to the dashboard (admins have richer admin pages).
    """
    role = getattr(current_user, 'role', None)
    if role not in (ROLE_STUDENT, ROLE_UNIVERSITY_COORD):
        flash('My University is only available to students and university coordinators.', 'info')
        return redirect(url_for('user.dashboard'))

    # Resolve university: coordinators via UniversityMember; students via user.university_id.
    univ = None
    if role == ROLE_UNIVERSITY_COORD:
        univ = _my_university()
    if univ is None and getattr(current_user, 'university_id', None):
        univ = db.session.get(University, current_user.university_id)

    if univ is None:
        flash('You are not linked to any university yet. Please contact your administrator.', 'warning')
        return render_template('university/my_university.html',
                               univ=None, can_edit=False, locked=False, is_student=(role == ROLE_STUDENT))

    is_coord = (role == ROLE_UNIVERSITY_COORD)
    locked = bool(getattr(univ, 'is_verified', False))
    can_edit = is_coord and not locked

    if request.method == 'POST':
        if not can_edit:
            if is_coord and locked:
                flash('This university has been verified — please contact an administrator for changes.', 'warning')
            else:
                abort(403)
            return redirect(url_for('university.my_university'))

        univ.name          = (request.form.get('name') or univ.name).strip()
        univ.description   = (request.form.get('description') or '').strip() or None
        univ.location      = (request.form.get('location') or '').strip() or None
        univ.website       = (request.form.get('website') or '').strip() or None
        univ.contact_email = (request.form.get('contact_email') or '').strip() or None
        univ.contact_phone = (request.form.get('contact_phone') or '').strip() or None

        logo = request.files.get('logo')
        if logo and logo.filename:
            try:
                univ.logo_filename = save_company_image(logo)
            except ValueError as e:
                flash(str(e), 'warning')

        banner = request.files.get('banner')
        if banner and banner.filename:
            try:
                univ.banner_filename = save_company_image(banner)
            except ValueError as e:
                flash(str(e), 'warning')

        try:
            log_audit('university.coordinator_edit', univ.name)
        except Exception:
            current_app.logger.exception('audit log failed for coordinator university edit')
        db.session.commit()
        flash('University profile updated.', 'success')
        return redirect(url_for('university.my_university'))

    return render_template('university/my_university.html',
                           univ=univ, can_edit=can_edit, locked=locked,
                           is_student=(role == ROLE_STUDENT))


@university_bp.route('/my-coordinator')
@login_required
def my_coordinator():
    """Student-facing page showing their university coordinator(s)."""
    if getattr(current_user, 'role', None) != ROLE_STUDENT:
        flash('My Coordinator is only available to students.', 'info')
        return redirect(url_for('user.dashboard'))

    univ = None
    if getattr(current_user, 'university_id', None):
        univ = db.session.get(University, current_user.university_id)

    coordinators = []
    if univ:
        members = (UniversityMember.query
                   .filter_by(university_id=univ.id)
                   .join(User, UniversityMember.user_id == User.id)
                   .filter(User.role == ROLE_UNIVERSITY_COORD, User.is_active == True)
                   .all())
        # Prefer coordinators in the student's department (if any), then others
        my_dept = current_user.university_department_id
        coordinators = sorted(members, key=lambda m: (
            0 if (my_dept and m.department_id == my_dept) else 1,
            (m.user.full_name or '').lower(),
        ))

    return render_template('university/my_coordinator.html',
                           univ=univ, coordinators=coordinators)


@university_bp.route('/profile/edit', methods=['GET', 'POST'])
@university_coordinator_required
def profile_edit():
    """Legacy alias — redirect to the unified My University page."""
    return redirect(url_for('university.my_university'), code=302)


@university_bp.route('/students')
@university_coordinator_required
def students():
    univ = _my_university()
    if not univ and current_user.role != ROLE_ADMIN:
        flash('You are not linked to any university.', 'warning')
        return redirect(url_for('university.dashboard'))

    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()

    membership = _my_membership()
    scoped_ids = _student_ids(univ, membership)
    if scoped_ids:
        q = User.query.filter(User.id.in_(scoped_ids), User.role == ROLE_STUDENT)
    else:
        q = User.query.filter(sql_false())

    if search:
        q = q.filter(or_(
            User.full_name.ilike(f'%{search}%'),
            User.email.ilike(f'%{search}%'),
            User.university_major.ilike(f'%{search}%'),
            User.university_class.ilike(f'%{search}%'),
        ))

    students_page = q.order_by(User.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False)

    return render_template('university/students.html',
        univ=univ, students=students_page, search=search)


@university_bp.route('/applications')
@university_coordinator_required
def applications():
    univ = _my_university()
    if not univ and current_user.role != ROLE_ADMIN:
        flash('You are not linked to any university.', 'warning')
        return redirect(url_for('university.dashboard'))

    page     = request.args.get('page', 1, type=int)
    search   = request.args.get('q', '').strip()
    status_f = request.args.get('status', '').strip()

    membership = _my_membership()
    student_ids = _student_ids(univ, membership)
    q = _internship_apps(student_ids)

    if search:
        q = q.join(Application.applicant).filter(or_(
            User.full_name.ilike(f'%{search}%'),
            User.email.ilike(f'%{search}%'),
        ))
    if status_f:
        q = q.filter(Application.status == status_f)

    apps = q.order_by(Application.applied_at.desc()).paginate(
        page=page, per_page=20, error_out=False)

    # KPI counts (unfiltered)
    base_q = _internship_apps(student_ids)
    kpi_total     = base_q.count()
    kpi_active    = base_q.filter(Application.status.in_(
        [STATUS_UNIV_PENDING, STATUS_NEW, STATUS_REVIEW, STATUS_INTERVIEW, STATUS_OFFER])).count()
    kpi_hired     = base_q.filter_by(status=STATUS_HIRED).count()
    kpi_rejected  = base_q.filter_by(status=STATUS_REJECTED).count()

    return render_template('university/applications.html',
        univ=univ, apps=apps, search=search, status_f=status_f,
        ALL_STATUSES=ALL_STATUSES,
        kpi_total=kpi_total, kpi_active=kpi_active,
        kpi_hired=kpi_hired, kpi_rejected=kpi_rejected)


@university_bp.route('/applications/<int:app_id>')
@university_coordinator_required
def application_detail(app_id):
    app = Application.query.get_or_404(app_id)
    univ = _my_university()
    membership = _my_membership()
    if univ:
        sids = _student_ids(univ, membership)
        if app.applicant_id not in sids:
            abort(403)
    if app.position.type != 'Internship':
        abort(403)
    return render_template('university/application_detail.html', app=app, univ=univ)


@university_bp.route('/applications/<int:app_id>/approve', methods=['POST'])
@university_coordinator_required
def application_approve(app_id):
    app = Application.query.get_or_404(app_id)
    univ = _my_university()
    if not univ:
        abort(403)
    scoped_student_ids = _student_ids(univ, _my_membership())
    if app.applicant_id not in scoped_student_ids or app.position.type != 'Internship':
        abort(403)
    if app.status != STATUS_UNIV_PENDING:
        flash('This application is not waiting for university approval.', 'warning')
        return redirect(url_for('university.applications'))

    app.status = STATUS_NEW
    db.session.add(ApplicationHistory(
        application_id=app.id,
        old_status=STATUS_UNIV_PENDING,
        new_status=STATUS_NEW,
        changed_by_id=current_user.id,
        note='Approved by university coordinator and forwarded to company team.',
    ))
    log_audit('university.app_approve', f'{app.applicant.full_name} -> {app.position.title}', user_id=current_user.id)
    db.session.commit()

    site_url = current_app.config['SITE_URL']

    # Notify applicant
    try:
        html = render_template('emails/application_received.html',
                               app=app, user=app.applicant, pos=app.position, site_url=site_url)
        send_email(app.applicant.email, f'Application forwarded to company review - {app.position.title}', html)
    except Exception as e:
        current_app.logger.warning(f'Applicant approval email failed: {e}')

    # Notify company staff (same flow as direct application)
    if app.position.company_id:
        manager_ids = [m.user_id for m in CompanyMember.query.filter_by(
            company_id=app.position.company_id, role='manager').all()]
        supervisors = User.query.filter(
            User.id.in_(manager_ids), User.is_active == True).all() if manager_ids else []
    else:
        supervisors = []

    for staff in supervisors:
        try:
            push_notification(
                staff.id,
                f'New approved internship application: {app.applicant.full_name} -> {app.position.title}',
                url_for('supervisor.application_detail', app_id=app.id),
            )
        except Exception:
            pass

    for staff in supervisors:
        review_url = site_url + url_for('supervisor.application_detail', app_id=app.id)
        try:
            html = render_template('emails/application_notify_staff.html',
                                   app=app, pos=app.position,
                                   recipient_name=staff.full_name.split()[0],
                                   review_url=review_url,
                                   coordinator_approved=True)
            send_email(staff.email,
                       f'Approved Internship Application: {app.applicant.full_name} -> {app.position.title}',
                       html)
        except Exception as e:
            current_app.logger.warning(f'Staff notification email to {staff.email} failed: {e}')

    flash('Application approved and forwarded to company review.', 'success')
    return redirect(url_for('university.applications'))


# ─── Student management ──────────────────────────────────────────────────────

@university_bp.route('/students/add', methods=['GET', 'POST'])
@university_coordinator_required
def student_add():
    univ = _my_university()
    membership = _my_membership()
    if not univ:
        flash('You are not linked to any university.', 'warning')
        return redirect(url_for('university.dashboard'))

    departments = UniversityDepartment.query.filter_by(university_id=univ.id, is_active=True).order_by(UniversityDepartment.name.asc()).all()

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        if not email:
            flash('Email is required.', 'danger')
            return render_template('university/student_form.html', univ=univ, student=None, departments=departments, scope_membership=membership)

        dept_id = request.form.get('university_department_id', type=int)
        class_scope = request.form.get('university_class', '').strip() or None

        if membership and membership.department_id:
            dept_id = membership.department_id
        if membership and membership.class_scope:
            class_scope = membership.class_scope

        if dept_id:
            dept = UniversityDepartment.query.filter_by(id=dept_id, university_id=univ.id).first()
            if not dept:
                flash('Invalid department selected.', 'danger')
                return render_template('university/student_form.html', univ=univ, student=None, departments=departments, scope_membership=membership)

        # Check if email exists — link existing user or create new one
        existing = User.query.filter_by(email=email).first()
        if existing:
            if existing.role != ROLE_STUDENT:
                flash(f'A non-student account already exists for {email}.', 'danger')
                return render_template('university/student_form.html', univ=univ, student=None, departments=departments, scope_membership=membership)
            existing.university_id = univ.id
            existing.university_name = univ.name
            existing.university_department_id = dept_id
            existing.university_class = class_scope
            existing.university_major = request.form.get('university_major', '').strip() or existing.university_major
            existing.student_gpa = request.form.get('student_gpa', '').strip() or existing.student_gpa
            existing.graduation_year = _parse_int(request.form.get('graduation_year', '')) or existing.graduation_year
            existing.student_id_number = request.form.get('student_id_number', '').strip() or existing.student_id_number
            log_audit('university.student_link', f'{existing.full_name} → {univ.name}', user_id=current_user.id)
            db.session.commit()
            flash(f'Existing student {existing.full_name} linked to {univ.name}.', 'success')
            return redirect(url_for('university.students'))

        full_name = request.form.get('full_name', '').strip()
        if not full_name:
            flash('Full name is required.', 'danger')
            return render_template('university/student_form.html', univ=univ, student=None, departments=departments, scope_membership=membership)

        temp_pw = secrets.token_urlsafe(10)
        student = User(
            full_name=full_name,
            email=email,
            phone=request.form.get('phone', '').strip() or None,
            role=ROLE_STUDENT,
            is_active=True,
            university_id=univ.id,
            university_name=univ.name,
            university_department_id=dept_id,
            university_class=class_scope,
            university_major=request.form.get('university_major', '').strip() or None,
            student_gpa=request.form.get('student_gpa', '').strip() or None,
            graduation_year=_parse_int(request.form.get('graduation_year', '')),
            student_id_number=request.form.get('student_id_number', '').strip() or None,
        )
        student.set_password(temp_pw)
        db.session.add(student)
        log_audit('university.student_add', f'{full_name} → {univ.name}', user_id=current_user.id)
        db.session.commit()

        try:
            html = render_template('emails/welcome_student.html',
                                   student=student, univ=univ, temp_password=temp_pw)
            send_email(student.email, 'Welcome to MOF Jobs — Your Student Account', html)
        except Exception as e:
            current_app.logger.warning(f'Student welcome email failed: {e}')

        flash(f'Student {full_name} added successfully.', 'success')
        return redirect(url_for('university.students'))

    return render_template('university/student_form.html', univ=univ, student=None, departments=departments, scope_membership=membership)


@university_bp.route('/students/<int:student_id>/edit', methods=['GET', 'POST'])
@university_coordinator_required
def student_edit(student_id):
    univ = _my_university()
    membership = _my_membership()
    student = db.get_or_404(User, student_id)
    departments = UniversityDepartment.query.filter_by(university_id=univ.id, is_active=True).order_by(UniversityDepartment.name.asc()).all() if univ else []

    scoped_student_ids = _student_ids(univ, membership)
    if univ and student.id not in scoped_student_ids:
        abort(403)

    if request.method == 'POST':
        student.full_name = request.form.get('full_name', student.full_name).strip()
        new_email = request.form.get('email', '').strip().lower()
        if new_email and new_email != student.email:
            clash = User.query.filter(User.email == new_email, User.id != student.id).first()
            if clash:
                flash('That email address is already in use by another account.', 'danger')
                return render_template('university/student_form.html', univ=univ, student=student, departments=departments, scope_membership=membership)
            student.email = new_email

        dept_id = request.form.get('university_department_id', type=int)
        class_scope = request.form.get('university_class', '').strip() or None
        if membership and membership.department_id:
            dept_id = membership.department_id
        if membership and membership.class_scope:
            class_scope = membership.class_scope
        if dept_id:
            dept = UniversityDepartment.query.filter_by(id=dept_id, university_id=univ.id).first()
            if not dept:
                flash('Invalid department selected.', 'danger')
                return render_template('university/student_form.html', univ=univ, student=student, departments=departments, scope_membership=membership)

        student.university_department_id = dept_id
        student.university_class = class_scope
        student.phone = request.form.get('phone', '').strip() or None
        student.headline = request.form.get('headline', '').strip() or None
        student.bio = request.form.get('bio', '').strip() or None
        student.location_city = request.form.get('location_city', '').strip() or None
        student.nationality = request.form.get('nationality', '').strip() or None
        student.gender = request.form.get('gender', '').strip() or None
        student.linkedin_url = request.form.get('linkedin_url', '').strip() or None
        student.github_url = request.form.get('github_url', '').strip() or None
        student.portfolio_url = request.form.get('portfolio_url', '').strip() or None
        student.university_major = request.form.get('university_major', '').strip() or None
        student.student_gpa = request.form.get('student_gpa', '').strip() or None
        student.graduation_year = _parse_int(request.form.get('graduation_year', ''))
        student.student_id_number = request.form.get('student_id_number', '').strip() or None
        new_pw = request.form.get('new_password', '').strip()
        if new_pw:
            student.set_password(new_pw)
        log_audit('university.student_edit', student.full_name, user_id=current_user.id)
        db.session.commit()
        flash('Student updated.', 'success')
        return redirect(url_for('university.students'))

    return render_template('university/student_form.html', univ=univ, student=student, departments=departments, scope_membership=membership)


@university_bp.route('/students/<int:student_id>/delete', methods=['POST'])
@university_coordinator_required
def student_delete(student_id):
    univ = _my_university()
    membership = _my_membership()
    student = db.get_or_404(User, student_id)
    if univ and student.id not in _student_ids(univ, membership):
        abort(403)
    name = student.full_name
    student.university_id = None
    student.university_department_id = None
    student.university_class = None
    student.university_name = None
    log_audit('university.student_remove', f'{name} ← {univ.name if univ else "university"}', user_id=current_user.id)
    db.session.commit()
    flash(f'Student {name} removed from your university.', 'success')
    return redirect(url_for('university.students'))


# ─── Internship application creation ─────────────────────────────────────────

@university_bp.route('/applications/new', methods=['GET', 'POST'])
@university_coordinator_required
def application_new():
    univ = _my_university()
    if not univ:
        flash('You are not linked to any university.', 'warning')
        return redirect(url_for('university.dashboard'))

    # Only show internship positions
    open_positions = (Position.query
                      .filter_by(is_active=True, type='Internship')
                      .order_by(Position.title).all())
    student_ids = _student_ids(univ, _my_membership())
    students = User.query.filter(User.id.in_(student_ids)).order_by(User.full_name).all() if student_ids else []

    # Pre-select a position when coordinator clicks "Apply for Student" on a company / job page.
    preselected_position_id = request.args.get('position_id', type=int)

    if request.method == 'POST':
        student_id  = request.form.get('student_id', type=int)
        position_id = request.form.get('position_id', type=int)

        if not student_id or not position_id:
            flash('Please select both a student and a position.', 'danger')
            return render_template('university/application_form.html',
                                   univ=univ, students=students, positions=open_positions)

        student  = User.query.get(student_id)
        position = Position.query.get(position_id)
        if not student or not position:
            flash('Invalid student or position.', 'danger')
            return redirect(url_for('university.application_new'))

        # Check student belongs to this coordinator scope
        if univ and student.id not in student_ids:
            abort(403)

        # Check not already applied
        existing = Application.query.filter_by(
            applicant_id=student_id, position_id=position_id).first()
        if existing:
            flash(f'{student.full_name} has already applied for {position.title}.', 'warning')
            return redirect(url_for('university.applications'))

        app = Application(
            applicant_id   = student_id,
            position_id    = position_id,
            cover_letter   = request.form.get('cover_letter', '').strip() or None,
            source         = 'University',
            status         = STATUS_NEW,
            internship_duration   = request.form.get('internship_duration', '').strip() or None,
            academic_credit_required = bool(request.form.get('academic_credit_required')),
        )
        db.session.add(app)
        log_audit('university.app_submit',
                  f'{student.full_name} → {position.title}', user_id=current_user.id)
        db.session.commit()
        flash(f'Application for {student.full_name} → {position.title} submitted.', 'success')
        return redirect(url_for('university.applications'))

    return render_template('university/application_form.html',
                           univ=univ, students=students, positions=open_positions,
                           preselected_position_id=preselected_position_id)


# ─── Student export / import ─────────────────────────────────────────────────

_STUDENT_COLS = [
    ('student_id_number', 'Student ID'),
    ('full_name',         'Full Name'),
    ('email',             'Email'),
    ('phone',             'Phone'),
    ('university_major',  'Major'),
    ('student_gpa',       'GPA'),
    ('graduation_year',   'Graduation Year'),
]


@university_bp.route('/students/export')
@university_coordinator_required
def students_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    univ = _my_university()
    membership = _my_membership()
    scoped_ids = _student_ids(univ, membership)
    rows = User.query.filter(User.id.in_(scoped_ids)).order_by(User.full_name).all() if scoped_ids else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    header_fill = PatternFill('solid', fgColor='1a5c38')
    header_font = Font(color='FFFFFF', bold=True)
    headers = [col[1] for col in _STUDENT_COLS]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)

    for ri, s in enumerate(rows, 2):
        for ci, (attr, _) in enumerate(_STUDENT_COLS, 1):
            ws.cell(row=ri, column=ci, value=getattr(s, attr, None))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = f"{univ.name.replace(' ','_')}_students.xlsx" if univ else 'students.xlsx'
    return send_file(buf, as_attachment=True, download_name=name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@university_bp.route('/students/import-template')
@university_coordinator_required
def students_import_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    header_fill = PatternFill('solid', fgColor='1a5c38')
    header_font = Font(color='FFFFFF', bold=True)
    example_data = {
        'student_id_number': '20231001',
        'full_name':         'Ahmed Al-Rashidi',
        'email':             'ahmed@example.com',
        'phone':             '+96894123456',
        'university_major':  'Computer Science',
        'student_gpa':       '3.75',
        'graduation_year':   '2025',
    }
    for ci, (attr, label) in enumerate(_STUDENT_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(label) + 4, 18)
        ws.cell(row=2, column=ci, value=example_data.get(attr, ''))

    # Mark example row in light gray
    from openpyxl.styles import PatternFill as PF
    ex_fill = PF('solid', fgColor='F3F4F6')
    for ci in range(1, len(_STUDENT_COLS) + 1):
        ws.cell(row=2, column=ci).fill = ex_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='students_import_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@university_bp.route('/students/import', methods=['POST'])
@university_coordinator_required
def students_import():
    import openpyxl
    univ = _my_university()
    membership = _my_membership()
    if not univ:
        flash('You are not linked to any university.', 'warning')
        return redirect(url_for('university.students'))

    f = request.files.get('file')
    if not f or not f.filename:
        flash('No file selected.', 'danger')
        return redirect(url_for('university.students'))
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('university.students'))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash('Could not read the uploaded file. Make sure it is a valid Excel workbook.', 'danger')
        return redirect(url_for('university.students'))

    headers = [str(cell.value or '').strip() for cell in ws[1]]
    label_to_attr = {col[1]: col[0] for col in _STUDENT_COLS}
    col_map = {}  # column index → attribute name
    for ci, h in enumerate(headers):
        if h in label_to_attr:
            col_map[ci] = label_to_attr[h]

    if 'email' not in col_map.values() or 'full_name' not in col_map.values():
        flash('Import failed: the file must have "Email" and "Full Name" columns.', 'danger')
        return redirect(url_for('university.students'))

    created = linked = skipped = 0
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue  # skip blank rows
        data = {}
        for ci, attr in col_map.items():
            val = row[ci] if ci < len(row) else None
            data[attr] = str(val).strip() if val is not None else ''

        email = data.get('email', '').lower()
        full_name = data.get('full_name', '').strip()
        if not email or not full_name:
            errors.append(f'Row skipped — missing email or name: {data}')
            skipped += 1
            continue

        existing = User.query.filter_by(email=email).first()
        if existing:
            if existing.role == ROLE_STUDENT:
                existing.university_id = univ.id
                existing.university_name = univ.name
                if membership and membership.department_id:
                    existing.university_department_id = membership.department_id
                if membership and membership.class_scope:
                    existing.university_class = membership.class_scope
                for attr in ('university_major', 'student_gpa', 'student_id_number', 'phone'):
                    if data.get(attr):
                        setattr(existing, attr, data[attr])
                if data.get('graduation_year'):
                    existing.graduation_year = _parse_int(data['graduation_year'])
                linked += 1
            else:
                errors.append(f'Non-student account exists for {email} — skipped')
                skipped += 1
            continue

        temp_pw = secrets.token_urlsafe(10)
        student = User(
            full_name         = full_name,
            email             = email,
            phone             = data.get('phone') or None,
            role              = ROLE_STUDENT,
            is_active         = True,
            university_id     = univ.id,
            university_department_id = membership.department_id if membership else None,
            university_name   = univ.name,
            university_class  = membership.class_scope if membership else None,
            university_major  = data.get('university_major') or None,
            student_gpa       = data.get('student_gpa') or None,
            graduation_year   = _parse_int(data.get('graduation_year', '')),
            student_id_number = data.get('student_id_number') or None,
        )
        student.set_password(temp_pw)
        db.session.add(student)
        db.session.flush()  # get student.id before commit

        try:
            html = render_template('emails/welcome_student.html',
                                   student=student, univ=univ, temp_password=temp_pw)
            send_email(student.email, 'Welcome to MOF Jobs — Your Student Account', html)
        except Exception as e:
            current_app.logger.warning(f'Welcome email failed for {email}: {e}')

        log_audit('university.student_import', f'{full_name} → {univ.name}', user_id=current_user.id)
        created += 1

    db.session.commit()

    parts = []
    if created:  parts.append(f'{created} created')
    if linked:   parts.append(f'{linked} linked')
    if skipped:  parts.append(f'{skipped} skipped')
    msg = 'Import complete: ' + ', '.join(parts) + '.' if parts else 'No rows processed.'
    flash(msg, 'success' if not errors else 'warning')
    for err in errors[:5]:
        flash(err, 'warning')
    return redirect(url_for('university.students'))


# ─── helpers ─────────────────────────────────────────────────────────────────

def _student_ids(univ, membership=None):
    q = User.query.filter_by(role=ROLE_STUDENT)
    if univ:
        q = q.filter_by(university_id=univ.id)
    if membership and membership.department_id:
        q = q.filter(User.university_department_id == membership.department_id)
    if membership and membership.class_scope:
        q = q.filter(User.university_class == membership.class_scope)
    return [u.id for u in q.all()]


def _internship_pos_ids():
    """Return all internship position IDs (avoids JOIN in paginated queries)."""
    return [p.id for p in Position.query.filter_by(type='Internship').all()]


def _internship_apps(student_ids):
    if not student_ids:
        return Application.query.filter(sql_false())
    pos_ids = _internship_pos_ids()
    if not pos_ids:
        return Application.query.filter(sql_false())
    return Application.query.filter(
        Application.applicant_id.in_(student_ids),
        Application.position_id.in_(pos_ids),
    )


def _parse_int(s):
    try:
        return int(str(s).strip()) if s and str(s).strip() else None
    except (ValueError, TypeError):
        return None


def _cohort_insights(univ, membership, student_ids, internship_q):
    """Compute cohort intelligence for the coordinator dashboard.

    Returns a dict with these keys (always present, may be empty):
      - placement_rate            : int 0..100 (% of students who got Hired)
      - interview_rate            : int 0..100 (% of apps that reached Interview+)
      - response_avg_days         : float | None (avg days between Applied and first status change)
      - apps_last_30              : int (apps in last 30 days)
      - apps_prev_30              : int (apps in days 30-60 ago)
      - active_students           : int (students with >=1 application)
      - inactive_students         : int (students with 0 apps)
      - at_risk                   : list[dict] of up to 8 students needing outreach
      - top_positions             : list[(title, count)] top 5 positions students applied to
      - top_companies             : list[(name, count)] top 5 hiring partners
      - dept_breakdown            : list[dict{name,students,apps,placed}] per department (if no scope)
      - pending_approvals         : int (STATUS_UNIV_PENDING count)
    """
    insights = {
        'placement_rate': 0, 'interview_rate': 0,
        'response_avg_days': None,
        'apps_last_30': 0, 'apps_prev_30': 0,
        'active_students': 0, 'inactive_students': 0,
        'at_risk': [], 'top_positions': [], 'top_companies': [],
        'dept_breakdown': [], 'pending_approvals': 0,
    }
    if not student_ids:
        return insights
    now = datetime.utcnow()
    cutoff_30 = now - timedelta(days=30)
    cutoff_60 = now - timedelta(days=60)

    # Placement / interview rates
    total_apps = internship_q.count()
    if total_apps:
        hired = internship_q.filter(Application.status == STATUS_HIRED).count()
        reached_interview = internship_q.filter(
            Application.status.in_([STATUS_INTERVIEW, STATUS_OFFER, STATUS_HIRED])
        ).count()
        # Placement rate = % of students with at least one Hired
        hired_student_ids = {a.applicant_id for a in
                             internship_q.filter(Application.status == STATUS_HIRED).all()}
        insights['placement_rate'] = round(len(hired_student_ids) * 100 / max(len(student_ids), 1))
        insights['interview_rate'] = round(reached_interview * 100 / max(total_apps, 1))

    insights['apps_last_30'] = internship_q.filter(Application.applied_at >= cutoff_30).count()
    insights['apps_prev_30'] = internship_q.filter(
        Application.applied_at >= cutoff_60, Application.applied_at < cutoff_30).count()
    insights['pending_approvals'] = internship_q.filter(
        Application.status == STATUS_UNIV_PENDING).count()

    # Avg time-to-response — first ApplicationHistory entry per app where status changed
    apps_with_response = (db.session.query(
            Application.id, Application.applied_at, func.min(ApplicationHistory.created_at))
        .join(ApplicationHistory, ApplicationHistory.application_id == Application.id)
        .filter(Application.applicant_id.in_(student_ids))
        .filter(ApplicationHistory.new_status.isnot(None))
        .group_by(Application.id, Application.applied_at)
        .all())
    if apps_with_response:
        diffs = []
        for _aid, applied_at, first_change in apps_with_response:
            if applied_at and first_change and first_change > applied_at:
                diffs.append((first_change - applied_at).total_seconds() / 86400.0)
        if diffs:
            insights['response_avg_days'] = round(sum(diffs) / len(diffs), 1)

    # Active vs inactive students
    applicant_ids = {row[0] for row in db.session.query(Application.applicant_id)
                     .filter(Application.applicant_id.in_(student_ids)).distinct().all()}
    insights['active_students']   = len(applicant_ids)
    insights['inactive_students'] = max(len(student_ids) - len(applicant_ids), 0)

    # At-risk students: no apps in 30 days OR profile strength < 40
    inactive_ids = [sid for sid in student_ids if sid not in applicant_ids]
    if inactive_ids:
        students = (User.query.filter(User.id.in_(inactive_ids))
                    .order_by(User.created_at.desc()).limit(8).all())
        for s in students:
            insights['at_risk'].append({
                'id': s.id, 'name': s.full_name, 'email': s.email,
                'reason': 'No applications submitted yet',
                'profile_pct': s.profile_strength,
                'major': s.university_major or '—',
            })

    # Top positions/companies
    pos_counts = (db.session.query(Position.title, func.count(Application.id).label('c'))
                  .join(Application, Application.position_id == Position.id)
                  .filter(Application.applicant_id.in_(student_ids))
                  .group_by(Position.title).order_by(func.count(Application.id).desc())
                  .limit(5).all())
    insights['top_positions'] = [(t, c) for (t, c) in pos_counts]

    from models import Company
    comp_counts = (db.session.query(Company.name, func.count(Application.id).label('c'))
                   .join(Position, Position.company_id == Company.id)
                   .join(Application, Application.position_id == Position.id)
                   .filter(Application.applicant_id.in_(student_ids))
                   .group_by(Company.name).order_by(func.count(Application.id).desc())
                   .limit(5).all())
    insights['top_companies'] = [(n, c) for (n, c) in comp_counts]

    # Department breakdown — only show if coordinator scope is full university
    if univ and not (membership and membership.department_id):
        for dept in UniversityDepartment.query.filter_by(university_id=univ.id, is_active=True).all():
            dept_student_ids = [u.id for u in
                                User.query.filter_by(university_department_id=dept.id, role=ROLE_STUDENT).all()]
            if not dept_student_ids:
                continue
            dapps = Application.query.filter(Application.applicant_id.in_(dept_student_ids))
            insights['dept_breakdown'].append({
                'name': dept.name,
                'students': len(dept_student_ids),
                'apps': dapps.count(),
                'placed': dapps.filter_by(status=STATUS_HIRED).count(),
            })
        insights['dept_breakdown'].sort(key=lambda d: -d['apps'])
        insights['dept_breakdown'] = insights['dept_breakdown'][:6]

    return insights
