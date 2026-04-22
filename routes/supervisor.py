from datetime import datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, current_app
from sqlalchemy import or_, and_
from models import (db, Application, ApplicationHistory, Position, User, Message,
                    Company, CompanyMember, CompanyFollow, CompanyPhoto,
                    UniversityMember,
                    UserSkill, UserExperience, UserEducation,
                    ALL_STATUSES, STATUS_UNIV_PENDING, ROLE_SUPERVISOR, ROLE_ADMIN, ROLE_STUDENT)
from helpers import supervisor_or_admin_required, log_history, save_company_image, push_notification, send_email, log_audit
from flask_login import current_user

supervisor_bp = Blueprint('supervisor', __name__)


@supervisor_bp.route('/')
@supervisor_or_admin_required
def dashboard():
    # Supervisors see only apps assigned to them
    my_apps = (Application.query
               .filter_by(assigned_to_id=current_user.id)
               .order_by(Application.updated_at.desc())
               .all())

    status_counts = {}
    for s in ALL_STATUSES:
        status_counts[s] = sum(1 for a in my_apps if a.status == s)

    pipeline = {s: [a for a in my_apps if a.status == s] for s in ALL_STATUSES}

    # Managed companies with open job counts
    memberships = (CompanyMember.query
                   .filter_by(user_id=current_user.id, role='manager')
                   .all())
    my_companies = [m.company for m in memberships if m.company and m.company.is_active]

    return render_template('supervisor/dashboard.html',
        my_apps=my_apps, status_counts=status_counts, pipeline=pipeline,
        ALL_STATUSES=ALL_STATUSES, my_companies=my_companies)


@supervisor_bp.route('/applications')
@supervisor_or_admin_required
def applications():
    page     = request.args.get('page', 1, type=int)
    status_f = request.args.getlist('status')   # multi-select
    q_str    = request.args.get('q', '').strip()

    managed_pos_ids = _get_managed_pos_ids()

    q = (Application.query
         .join(User, Application.applicant_id == User.id)
         .filter(or_(
             Application.position_id.in_(managed_pos_ids),
             Application.assigned_to_id == current_user.id
         )))

    if status_f:
        q = q.filter(Application.status.in_(status_f))
    if q_str:
        q = q.filter(or_(User.full_name.ilike(f'%{q_str}%'),
                         User.email.ilike(f'%{q_str}%')))

    apps = q.order_by(Application.applied_at.desc()).paginate(page=page, per_page=20)
    return render_template('supervisor/applications.html',
        apps=apps, ALL_STATUSES=ALL_STATUSES, status_f=status_f, q_str=q_str)


@supervisor_bp.route('/applications/export')
@supervisor_or_admin_required
def applications_export():
    import csv, io
    from flask import Response
    from models import UserSkill, UserExperience, UserEducation
    fmt      = request.args.get('fmt', 'csv')
    status_f = request.args.getlist('status')
    q_str    = request.args.get('q', '').strip()

    managed_pos_ids = _get_managed_pos_ids()
    q = (Application.query
         .join(User, Application.applicant_id == User.id)
         .filter(or_(
             Application.position_id.in_(managed_pos_ids),
             Application.assigned_to_id == current_user.id
         )))
    if status_f:
        q = q.filter(Application.status.in_(status_f))
    if q_str:
        q = q.filter(or_(User.full_name.ilike(f'%{q_str}%'),
                         User.email.ilike(f'%{q_str}%')))
    apps = q.order_by(Application.applied_at.desc()).all()

    def fmt_date(d):
        return d.strftime('%Y-%m-%d') if d else ''

    def fmt_dt(d):
        return d.strftime('%Y-%m-%d %H:%M') if d else ''

    col_headers = [
        # ── Application ──────────────────────────────────────────────
        'App ID', 'Status', 'Source', 'Applied At', 'Last Updated',
        'Expected Salary', 'Assigned To', 'Internal Notes', 'Cover Letter', 'CV File',
        # ── Applicant ────────────────────────────────────────────────
        'Full Name', 'Email', 'Phone', 'Location', 'Date of Birth',
        'Nationality', 'Gender', 'Headline', 'Resume Summary',
        'LinkedIn', 'GitHub', 'Portfolio', 'Profile Joined',
        # ── Position / Company ───────────────────────────────────────
        'Position', 'Department', 'Position Location', 'Position Type',
        'Posted Salary Range', 'Company',
        # ── Profile data (semicolon-separated within each cell) ──────
        'Skills', 'Experience', 'Education', 'Languages', 'Certifications',
    ]

    rows = []
    for a in apps:
        u = a.applicant
        pos = a.position
        comp = pos.company if pos else None

        skills = '; '.join(
            f"{s.name} ({s.proficiency})" for s in
            UserSkill.query.filter_by(user_id=u.id).order_by(UserSkill.name).all()
        )
        experiences = '; '.join(
            f"{e.title} at {e.company or '?'} ({fmt_date(e.start_date)}–{fmt_date(e.end_date) or 'present'})"
            for e in UserExperience.query.filter_by(user_id=u.id)
                                        .order_by(UserExperience.start_date.desc().nullslast()).all()
        )
        educations = '; '.join(
            f"{ed.degree or ''} {ed.field or ''} — {ed.institution} ({ed.end_year or ''})"
            for ed in UserEducation.query.filter_by(user_id=u.id)
                                        .order_by(UserEducation.end_year.desc().nullslast()).all()
        )
        languages = ''
        certifications = ''

        rows.append([
            # Application
            a.id,
            a.status,
            a.source or '',
            fmt_dt(a.applied_at),
            fmt_dt(a.updated_at),
            a.expected_salary or '',
            a.assigned_to.full_name if a.assigned_to else '',
            a.notes or '',
            a.cover_letter or '',
            a.cv_original or a.cv_filename or '',
            # Applicant
            u.full_name,
            u.email,
            u.phone or '',
            u.location_city or '',
            fmt_date(u.date_of_birth),
            u.nationality or '',
            u.gender or '',
            u.headline or '',
            u.resume_headline or '',
            u.linkedin_url or '',
            u.github_url or '',
            u.portfolio_url or '',
            fmt_date(u.created_at),
            # Position / Company
            pos.title if pos else '',
            pos.department or '' if pos else '',
            pos.location or '' if pos else '',
            pos.type or '' if pos else '',
            pos.salary_range or '' if pos else '',
            comp.name if comp else '',
            # Profile
            skills,
            experiences,
            educations,
            languages,
            certifications,
        ])

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Section header colours
        SECTION_FILLS = {
            'application': PatternFill('solid', fgColor='198754'),   # green
            'applicant':   PatternFill('solid', fgColor='0D6EFD'),   # blue
            'position':    PatternFill('solid', fgColor='6C757D'),   # grey
            'profile':     PatternFill('solid', fgColor='7B3F9E'),   # purple
        }
        section_ranges = {
            'application': range(0, 10),
            'applicant':   range(10, 23),
            'position':    range(23, 29),
            'profile':     range(29, 34),
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Applications'
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, h in enumerate(col_headers, 1):
            section = next(
                (sec for sec, rng in section_ranges.items() if (col_idx - 1) in rng),
                'application'
            )
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = SECTION_FILLS[section]
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        for row in rows:
            ws.append(row)

        # Wrap long-text columns (cover letter, notes, skills, experience, etc.)
        wrap_cols = {col_headers.index(h) + 1 for h in
                     ('Cover Letter', 'Internal Notes', 'Skills', 'Experience',
                      'Education', 'Languages', 'Certifications')}
        for col_idx in wrap_cols:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx,
                                     min_row=2, max_row=ws.max_row):
                for c in cell:
                    c.alignment = Alignment(wrap_text=True, vertical='top')

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            col_idx = col[0].column
            if col_idx in wrap_cols:
                ws.column_dimensions[col_letter].width = 40
            else:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        ws.freeze_panes = 'A2'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="applications.xlsx"'}
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(col_headers)
    for row in rows:
        writer.writerow(row)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename="applications.csv"'}
    )


@supervisor_bp.route('/applications/<int:app_id>')
@supervisor_or_admin_required
def application_detail(app_id):
    app = db.get_or_404(Application, app_id)
    # Supervisors can view apps assigned to them OR for positions in their managed companies
    if current_user.role != 'admin':
        managed_pos_ids = _get_managed_pos_ids()
        if app.assigned_to_id != current_user.id and app.position_id not in managed_pos_ids:
            flash('You do not have access to this application.', 'danger')
            return redirect(url_for('supervisor.applications'))

    # Supervisors see all status changes + their own entries (notes without status change)
    history = app.history.filter(
        db.or_(
            ApplicationHistory.new_status.isnot(None),
            ApplicationHistory.changed_by_id == current_user.id
        )
    ).order_by(ApplicationHistory.created_at.desc()).all()
    interviews = app.interviews.order_by('scheduled_at').all()

    # Applicant profile
    u = app.applicant
    skills         = UserSkill.query.filter_by(user_id=u.id).order_by(UserSkill.name).all()
    experiences    = UserExperience.query.filter_by(user_id=u.id).order_by(
                        UserExperience.start_date.desc().nullslast()).all()
    educations     = UserEducation.query.filter_by(user_id=u.id).order_by(
                        UserEducation.end_year.desc().nullslast()).all()
    # Removed languages and certifications (rollback)

    # Message thread between this supervisor and the applicant
    thread_messages = (Message.query
                       .filter(or_(
                           and_(Message.sender_id == current_user.id,
                                Message.receiver_id == u.id),
                           and_(Message.sender_id == u.id,
                                Message.receiver_id == current_user.id)
                       ))
                       .order_by(Message.created_at.asc())
                       .all())
    Message.query.filter_by(sender_id=u.id, receiver_id=current_user.id,
                            is_read=False).update({'is_read': True})

    # Admin users + thread between this supervisor and any admin
    admin_users = User.query.filter_by(role=ROLE_ADMIN, is_active=True).all()
    admin_ids   = [a.id for a in admin_users]
    admin_thread = []
    if admin_ids:
        admin_thread = (Message.query
                        .filter(or_(
                            and_(Message.sender_id == current_user.id,
                                 Message.receiver_id.in_(admin_ids)),
                            and_(Message.sender_id.in_(admin_ids),
                                 Message.receiver_id == current_user.id)
                        ))
                        .order_by(Message.created_at.asc())
                        .all())
        Message.query.filter(
            Message.sender_id.in_(admin_ids),
            Message.receiver_id == current_user.id,
            Message.is_read == False
        ).update({'is_read': True}, synchronize_session=False)

    db.session.commit()

    available_statuses = ALL_STATUSES if (app.position and app.position.type == 'Internship') else [s for s in ALL_STATUSES if s != STATUS_UNIV_PENDING]
    return render_template('supervisor/application_detail.html',
        app=app, ALL_STATUSES=available_statuses, history=history, interviews=interviews,
        skills=skills, experiences=experiences, educations=educations,
        # Removed languages and certifications (rollback)
        thread_messages=thread_messages,
        admin_users=admin_users, admin_thread=admin_thread)


@supervisor_bp.route('/applications/<int:app_id>/update', methods=['POST'])
@supervisor_or_admin_required
def application_update(app_id):
    application = db.get_or_404(Application, app_id)

    if current_user.role != 'admin':
        managed_pos_ids = _get_managed_pos_ids()
        if (application.assigned_to_id != current_user.id
                and application.position_id not in managed_pos_ids):
            flash('Not authorized.', 'danger')
            return redirect(url_for('supervisor.applications'))

    new_status = request.form.get('status')
    note       = request.form.get('note', '').strip()

    if new_status and new_status != application.status:
        log_history(application, current_user, new_status=new_status, note=note or None, is_internal=True)
        application.status     = new_status
        application.updated_at = datetime.utcnow()
        log_audit('supervisor.status_change',
                  f'#{application.id} {application.position.title} → {new_status}',
                  user_id=current_user.id)
        db.session.commit()
        _send_status_email(application)
        push_notification(
            application.applicant_id,
            f'Your application for "{application.position.title}" is now: {new_status}',
            url_for('user.my_applications')
        )
    else:
        if note:
            log_history(application, current_user, note=note, is_internal=True)
        db.session.commit()
    flash('Application updated.', 'success')
    return redirect(url_for('supervisor.application_detail', app_id=app_id))


@supervisor_bp.route('/applications/<int:app_id>/delete', methods=['POST'])
@supervisor_or_admin_required
def application_delete(app_id):
    application = db.get_or_404(Application, app_id)
    if current_user.role != 'admin':
        managed_pos_ids = _get_managed_pos_ids()
        if (application.assigned_to_id != current_user.id
                and application.position_id not in managed_pos_ids):
            abort(403)
    application.history.delete(synchronize_session=False)
    application.interviews.delete(synchronize_session=False)
    db.session.delete(application)
    log_audit('supervisor.application_delete', f'#{app_id}', user_id=current_user.id)
    db.session.commit()
    flash('Application permanently deleted.', 'success')
    return redirect(url_for('supervisor.applications'))


@supervisor_bp.route('/applications/<int:app_id>/history/<int:history_id>/delete', methods=['POST'])
@supervisor_or_admin_required
def history_delete(app_id, history_id):
    entry = db.get_or_404(ApplicationHistory, history_id)
    if entry.application_id != app_id:
        abort(404)
    # Supervisors can only delete their own entries
    if current_user.role != 'admin' and entry.changed_by_id != current_user.id:
        abort(403)
    db.session.delete(entry)
    db.session.commit()
    flash('Timeline entry deleted.', 'success')
    return redirect(url_for('supervisor.application_detail', app_id=app_id))


# ─── POSITION MANAGEMENT ──────────────────────────────────────────────────────

@supervisor_bp.route('/positions')
@supervisor_or_admin_required
def positions():
    page     = request.args.get('page', 1, type=int)
    q_str    = request.args.get('q', '').strip()
    types_f  = request.args.getlist('type')
    status_f = request.args.get('status', '')

    managed_company_ids = _get_managed_company_ids()
    if managed_company_ids:
        q = Position.query.filter(Position.company_id.in_(managed_company_ids))
    else:
        q = Position.query.filter(db.false())

    if q_str:
        q = q.filter(or_(Position.title.ilike(f'%{q_str}%'),
                         Position.department.ilike(f'%{q_str}%')))
    if types_f:
        q = q.filter(Position.type.in_(types_f))
    if status_f == 'active':
        q = q.filter_by(is_active=True)
    elif status_f == 'inactive':
        q = q.filter_by(is_active=False)

    positions = q.order_by(Position.created_at.desc()).paginate(page=page, per_page=25)
    managed_companies = [m.company for m in
                         CompanyMember.query.filter_by(user_id=current_user.id, role='manager').all()
                         if m.company]

    return render_template('supervisor/positions.html',
                           positions=positions, q_str=q_str, types_f=types_f,
                           status_f=status_f, managed_companies=managed_companies,
                           POSITION_TYPES=POSITION_TYPES)


@supervisor_bp.route('/positions/<int:pos_id>/toggle', methods=['POST'])
@supervisor_or_admin_required
def position_toggle(pos_id):
    pos = db.get_or_404(Position, pos_id)
    if pos.company_id not in _get_managed_company_ids():
        abort(403)
    pos.is_active = not pos.is_active
    db.session.commit()
    state = 'activated' if pos.is_active else 'deactivated'
    flash(f'"{pos.title}" {state}.', 'success')
    return redirect(request.referrer or url_for('supervisor.positions'))


@supervisor_bp.route('/positions/<int:pos_id>/delete', methods=['POST'])
@supervisor_or_admin_required
def position_delete(pos_id):
    pos = db.get_or_404(Position, pos_id)
    if pos.company_id not in _get_managed_company_ids():
        abort(403)
    title = pos.title
    app_count = pos.application_count
    for app in pos.applications.all():
        app.history.delete(synchronize_session=False)
        app.interviews.delete(synchronize_session=False)
        db.session.delete(app)
    db.session.delete(pos)
    db.session.commit()
    flash(f'Position "{title}" and {app_count} application(s) permanently deleted.', 'success')
    return redirect(url_for('supervisor.positions'))


# ─── SHARED CONSTANTS ────────────────────────────────────────────────────────

def _send_status_email(application):
    """Send applicant a status update email when supervisor changes status."""
    try:
        html = render_template('emails/status_update.html', app=application)
        send_email(application.applicant.email,
                   f'Your application update — {application.position.title}',
                   html)
    except Exception as e:
        current_app.logger.error(f'Status email failed for app #{application.id}: {e}')
        flash(f'Status updated, but email to {application.applicant.email} could not be sent: {e}', 'warning')

INDUSTRY_CHOICES = [
    'Engineering', 'Construction', 'Oil & Gas', 'Manufacturing',
    'Information Technology', 'Finance', 'Healthcare', 'Education',
    'Consulting', 'Government', 'Logistics', 'Real Estate', 'Other',
]
SIZE_CHOICES = ['1–10', '11–50', '51–200', '201–500', '500+']
POSITION_TYPES = ['Full-time', 'Part-time', 'Contract', 'Internship', 'Remote', 'Freelance']


# ─── COMPANY MANAGEMENT (supervisor as manager) ───────────────────────────────

def _get_managed_company_ids():
    """Return list of company IDs managed by current user."""
    return [m.company_id for m in
            CompanyMember.query.filter_by(user_id=current_user.id, role='manager').all()]


def _get_managed_pos_ids():
    """Return list of position IDs in companies managed by current user."""
    cids = _get_managed_company_ids()
    if not cids:
        return []
    return [p.id for p in Position.query.filter(Position.company_id.in_(cids)).all()]


def _get_managed_company_or_403(company_id):
    """Return company if current supervisor manages it, else 403."""
    member = CompanyMember.query.filter_by(
        company_id=company_id,
        user_id=current_user.id,
        role='manager',
    ).first()
    if not member:
        abort(403)
    return db.get_or_404(Company, company_id)


@supervisor_bp.route('/companies')
@supervisor_or_admin_required
def companies():
    managed = (CompanyMember.query
               .filter_by(user_id=current_user.id, role='manager')
               .all())
    total_jobs  = sum(m.company.open_jobs_count for m in managed if m.company)
    total_apps  = sum(
        Application.query.filter(
            Application.position_id.in_(
                [p.id for p in m.company.positions] if m.company else []
            )
        ).count()
        for m in managed
    )
    return render_template('supervisor/companies.html',
                           memberships=managed,
                           kpi_companies=len(managed),
                           kpi_jobs=total_jobs,
                           kpi_apps=total_apps)


@supervisor_bp.route('/companies/<int:company_id>')
@supervisor_or_admin_required
def company_dashboard(company_id):
    company   = _get_managed_company_or_403(company_id)
    positions = company.positions.order_by(Position.created_at.desc()).all()
    active_positions = [p for p in positions if p.is_active]
    total_apps  = sum(p.application_count for p in positions)
    total_followers = company.follower_count
    recent_apps = (Application.query
                   .join(Position)
                   .filter(Position.company_id == company_id)
                   .order_by(Application.applied_at.desc())
                   .limit(5).all())
    photos = company.photos.all()
    followers = (CompanyFollow.query
                 .filter_by(company_id=company_id)
                 .order_by(CompanyFollow.followed_at.desc())
                 .all())
    return render_template('supervisor/company.html',
                           company=company, positions=positions,
                           active_positions=active_positions,
                           total_apps=total_apps,
                           total_followers=total_followers,
                           recent_apps=recent_apps,
                           photos=photos,
                           followers=followers)


@supervisor_bp.route('/companies/<int:company_id>/edit', methods=['GET', 'POST'])
@supervisor_or_admin_required
def company_edit(company_id):
    company = _get_managed_company_or_403(company_id)
    # Verified + active companies are locked for supervisors
    if company.is_verified and company.is_active and current_user.role != 'admin':
        flash('This company is verified and active. Only an administrator can edit its information.', 'warning')
        return redirect(url_for('supervisor.company_dashboard', company_id=company_id))

    if request.method == 'GET':
        return render_template('admin/company_form.html', company=company,
                               industries=INDUSTRY_CHOICES, sizes=SIZE_CHOICES,
                               form_action=url_for('supervisor.company_edit',
                                                   company_id=company_id),
                               cancel_url=url_for('supervisor.companies'))

    company.description   = request.form.get('description', '').strip()
    company.industry      = request.form.get('industry', company.industry)
    company.size          = request.form.get('size', company.size)
    company.website       = request.form.get('website', '').strip()
    company.location      = request.form.get('location', '').strip()
    company.contact_email = request.form.get('contact_email', '').strip()
    company.contact_phone = request.form.get('contact_phone', '').strip()
    yr = request.form.get('founded_year', '').strip()
    company.founded_year  = int(yr) if yr.isdigit() else company.founded_year

    logo = request.files.get('logo')
    if logo and logo.filename:
        try:
            company.logo_filename = save_company_image(logo)
        except ValueError as e:
            flash(str(e), 'danger')
    cover = request.files.get('cover')
    if cover and cover.filename:
        try:
            company.cover_filename = save_company_image(cover)
        except ValueError as e:
            flash(str(e), 'danger')

    db.session.commit()
    flash('Company information updated.', 'success')
    return redirect(url_for('supervisor.companies'))


@supervisor_bp.route('/companies/<int:company_id>/delete', methods=['POST'])
@supervisor_or_admin_required
def company_delete(company_id):
    company = _get_managed_company_or_403(company_id)
    name = company.name
    db.session.delete(company)
    db.session.commit()
    flash(f'Company "{name}" permanently deleted.', 'success')
    return redirect(url_for('supervisor.companies'))


# ── Company photo gallery ────────────────────────────────────────────────────

@supervisor_bp.route('/companies/<int:company_id>/photos/upload', methods=['POST'])
@supervisor_or_admin_required
def company_photo_upload(company_id):
    import os, uuid as _uuid
    from werkzeug.utils import secure_filename
    from flask import current_app
    company = _get_managed_company_or_403(company_id)

    existing = company.photos.count()
    if existing >= 10:
        flash('Maximum of 10 photos allowed. Delete one first.', 'warning')
        return redirect(url_for('supervisor.company_dashboard', company_id=company_id) + '#tab-gallery')

    files = request.files.getlist('photos')
    allowed_exts = {'jpg', 'jpeg', 'png', 'webp', 'gif'}
    added = 0
    for f in files:
        if not f or not f.filename:
            continue
        if existing + added >= 10:
            flash('Some photos were skipped — 10-photo limit reached.', 'warning')
            break
        ext = secure_filename(f.filename).rsplit('.', 1)[-1].lower()
        if ext not in allowed_exts:
            flash(f'Skipped {f.filename} — unsupported format.', 'warning')
            continue
        stored = f"{_uuid.uuid4().hex}.{ext}"
        folder = current_app.config['COMPANY_PHOTOS_FOLDER']
        os.makedirs(folder, exist_ok=True)
        f.save(os.path.join(folder, stored))
        photo = CompanyPhoto(company_id=company_id, filename=stored,
                             caption=request.form.get('caption', '').strip() or None,
                             uploaded_by=current_user.id)
        db.session.add(photo)
        added += 1

    db.session.commit()
    if added:
        flash(f'{added} photo{"s" if added > 1 else ""} uploaded.', 'success')
    return redirect(url_for('supervisor.company_dashboard', company_id=company_id) + '#tab-gallery')


@supervisor_bp.route('/companies/<int:company_id>/photos/<int:photo_id>/delete', methods=['POST'])
@supervisor_or_admin_required
def company_photo_delete(company_id, photo_id):
    import os
    from flask import current_app
    _get_managed_company_or_403(company_id)
    photo = db.get_or_404(CompanyPhoto, photo_id)
    if photo.company_id != company_id:
        abort(403)
    try:
        path = os.path.join(current_app.config['COMPANY_PHOTOS_FOLDER'], photo.filename)
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass
    db.session.delete(photo)
    db.session.commit()
    flash('Photo removed.', 'success')
    return redirect(url_for('supervisor.company_dashboard', company_id=company_id) + '#tab-gallery')


# ── Follower email blast ─────────────────────────────────────────────────────

@supervisor_bp.route('/companies/<int:company_id>/followers/email', methods=['POST'])
@supervisor_or_admin_required
def followers_email(company_id):
    from helpers import send_email
    _get_managed_company_or_403(company_id)
    user_ids  = request.form.getlist('user_ids')
    subject   = request.form.get('subject', '').strip()
    body_html = request.form.get('body', '').strip()

    if not user_ids:
        flash('No followers selected.', 'warning')
        return redirect(url_for('supervisor.company_dashboard', company_id=company_id) + '#tab-followers')
    if not subject or not body_html:
        flash('Subject and message body are required.', 'warning')
        return redirect(url_for('supervisor.company_dashboard', company_id=company_id) + '#tab-followers')

    recipients = User.query.filter(User.id.in_(user_ids)).all()
    sent = 0
    for user in recipients:
        try:
            send_email(user.email, subject, body_html)
            sent += 1
        except Exception:
            pass

    flash(f'Email sent to {sent} follower{"s" if sent != 1 else ""}.', 'success')
    return redirect(url_for('supervisor.company_dashboard', company_id=company_id) + '#tab-followers')


# ── Follower export ──────────────────────────────────────────────────────────

@supervisor_bp.route('/companies/<int:company_id>/followers/export')
@supervisor_or_admin_required
def followers_export(company_id):
    import csv, io
    from flask import Response
    company = _get_managed_company_or_403(company_id)
    fmt = request.args.get('fmt', 'csv')

    follows = (CompanyFollow.query
               .filter_by(company_id=company_id)
               .order_by(CompanyFollow.followed_at.desc())
               .all())

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Followers'

        headers = ['Full Name', 'Email', 'Phone', 'Location', 'Headline', 'Role', 'Followed At']
        header_fill = PatternFill('solid', fgColor='198754')
        header_font = Font(bold=True, color='FFFFFF')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, f in enumerate(follows, 2):
            u = f.user
            ws.append([
                u.full_name,
                u.email,
                u.phone or '',
                u.location_city or '',
                u.headline or u.resume_headline or '',
                u.role,
                f.followed_at.strftime('%Y-%m-%d %H:%M') if f.followed_at else '',
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'{company.slug}-followers.xlsx'
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )

    # Default: CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Full Name', 'Email', 'Phone', 'Location', 'Headline', 'Role', 'Followed At'])
    for f in follows:
        u = f.user
        writer.writerow([
            u.full_name,
            u.email,
            u.phone or '',
            u.location_city or '',
            u.headline or u.resume_headline or '',
            u.role,
            f.followed_at.strftime('%Y-%m-%d %H:%M') if f.followed_at else '',
        ])
    filename = f'{company.slug}-followers.csv'
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@supervisor_bp.route('/companies/<int:company_id>/jobs/new', methods=['GET', 'POST'])
@supervisor_or_admin_required
def company_job_new(company_id):
    company = _get_managed_company_or_403(company_id)
    if request.method == 'POST':
        is_active = bool(request.form.get('is_active'))
        pos = Position(
            title        = request.form['title'].strip(),
            department   = request.form.get('department', '').strip(),
            location     = request.form.get('location', company.location or 'Slemani, Iraq').strip(),
            type         = request.form.get('type', 'Full-time'),
            description  = request.form.get('description', '').strip(),
            requirements = request.form.get('requirements', '').strip(),
            salary_range = request.form.get('salary_range', '').strip(),
            is_active    = is_active,
            company_id   = company.id,
            created_by   = current_user.id,
        )
        closes = request.form.get('closes_at', '').strip()
        if closes:
            pos.closes_at = datetime.strptime(closes, '%Y-%m-%d')
        db.session.add(pos)
        db.session.commit()

        if is_active:
            _send_company_job_alerts(pos)

        flash(f'Job "{pos.title}" posted for {company.name}.', 'success')
        return redirect(url_for('supervisor.company_dashboard', company_id=company_id))
    return render_template('supervisor/company_job_form.html',
                           company=company, pos=None)


@supervisor_bp.route('/companies/<int:company_id>/jobs/<int:pos_id>/edit',
                     methods=['GET', 'POST'])
@supervisor_or_admin_required
def company_job_edit(company_id, pos_id):
    company = _get_managed_company_or_403(company_id)
    pos = db.get_or_404(Position, pos_id)
    if pos.company_id != company_id:
        abort(403)

    if request.method == 'POST':
        was_active = pos.is_active
        pos.title        = request.form['title'].strip()
        pos.department   = request.form.get('department', '').strip()
        pos.location     = request.form.get('location', '').strip()
        pos.type         = request.form.get('type', 'Full-time')
        pos.description  = request.form.get('description', '').strip()
        pos.requirements = request.form.get('requirements', '').strip()
        pos.salary_range = request.form.get('salary_range', '').strip()
        pos.is_active    = bool(request.form.get('is_active'))
        closes = request.form.get('closes_at', '').strip()
        pos.closes_at    = datetime.strptime(closes, '%Y-%m-%d') if closes else None
        db.session.commit()

        if pos.is_active and not was_active:
            _send_company_job_alerts(pos)

        flash('Job updated.', 'success')
        return redirect(url_for('supervisor.company_dashboard', company_id=company_id))
    return render_template('supervisor/company_job_form.html',
                           company=company, pos=pos)


# ─── USERS (view-only — applicants for managed positions) ────────────────────

@supervisor_bp.route('/users')
@supervisor_or_admin_required
def users():
    q_str = request.args.get('q', '').strip()
    pos_f = request.args.get('position_id', 0, type=int)

    managed_pos_ids = _get_managed_pos_ids()
    if managed_pos_ids:
        q = (User.query
             .join(Application, Application.applicant_id == User.id)
             .filter(Application.position_id.in_(managed_pos_ids))
             .distinct())
        if pos_f and pos_f in managed_pos_ids:
            q = (User.query
                 .join(Application, Application.applicant_id == User.id)
                 .filter(Application.position_id == pos_f)
                 .distinct())
    else:
        q = User.query.filter(db.false())

    if q_str:
        q = q.filter(or_(User.full_name.ilike(f'%{q_str}%'),
                         User.email.ilike(f'%{q_str}%')))

    applicants = q.order_by(User.full_name).all()
    managed_positions = Position.query.filter(
        Position.id.in_(managed_pos_ids)).order_by(Position.title).all() \
        if managed_pos_ids else []

    return render_template('supervisor/users.html',
                           applicants=applicants, q_str=q_str, pos_f=pos_f,
                           managed_positions=managed_positions)


@supervisor_bp.route('/users/<int:user_id>')
@supervisor_or_admin_required
def user_detail(user_id):
    from models import UserSkill, UserExperience, UserEducation
    # Verify this user applied to at least one managed position (or is assigned to current supervisor)
    managed_pos_ids = _get_managed_pos_ids()
    user = db.get_or_404(User, user_id)
    has_access = Application.query.filter(
        Application.applicant_id == user_id,
        or_(
            Application.position_id.in_(managed_pos_ids),
            Application.assigned_to_id == current_user.id,
        )
    ).first()
    if not has_access and current_user.role != 'admin':
        flash('You do not have access to this applicant profile.', 'danger')
        return redirect(url_for('supervisor.users'))

    skills         = UserSkill.query.filter_by(user_id=user.id).order_by(UserSkill.name).all()
    experiences    = UserExperience.query.filter_by(user_id=user.id).order_by(
                        UserExperience.start_date.desc().nullslast()).all()
    educations     = UserEducation.query.filter_by(user_id=user.id).order_by(
                        UserEducation.end_year.desc().nullslast()).all()
    # Removed languages and certifications (rollback)

    # Only show applications to managed positions (or assigned to current supervisor)
    applications = Application.query.filter(
        Application.applicant_id == user_id,
        or_(
            Application.position_id.in_(managed_pos_ids),
            Application.assigned_to_id == current_user.id,
        )
    ).order_by(Application.applied_at.desc()).all()

    # Message thread
    thread_messages = (Message.query
                       .filter(or_(
                           and_(Message.sender_id == current_user.id,
                                Message.receiver_id == user.id),
                           and_(Message.sender_id == user.id,
                                Message.receiver_id == current_user.id)
                       ))
                       .order_by(Message.created_at.asc())
                       .all())
    Message.query.filter_by(sender_id=user.id, receiver_id=current_user.id,
                            is_read=False).update({'is_read': True})
    db.session.commit()

    return render_template('supervisor/user_detail.html',
                           user=user, skills=skills, experiences=experiences,
                           educations=educations, applications=applications,
                           thread_messages=thread_messages)


@supervisor_bp.route('/users/export')
@supervisor_or_admin_required
def users_export():
    import csv, io
    from flask import Response
    fmt   = request.args.get('fmt', 'csv')
    q_str = request.args.get('q', '').strip()
    pos_f = request.args.get('position_id', 0, type=int)

    managed_pos_ids = _get_managed_pos_ids()
    if managed_pos_ids:
        q = (User.query
             .join(Application, Application.applicant_id == User.id)
             .filter(Application.position_id.in_(managed_pos_ids))
             .distinct())
        if pos_f and pos_f in managed_pos_ids:
            q = (User.query
                 .join(Application, Application.applicant_id == User.id)
                 .filter(Application.position_id == pos_f)
                 .distinct())
    else:
        q = User.query.filter(db.false())
    if q_str:
        q = q.filter(or_(User.full_name.ilike(f'%{q_str}%'),
                         User.email.ilike(f'%{q_str}%')))
    applicants = q.order_by(User.full_name).all()

    col_headers = ['Full Name', 'Email', 'Phone', 'Location', 'Headline', 'Applications', 'Joined']
    rows = [[
        u.full_name,
        u.email,
        u.phone or '',
        u.location_city or '',
        u.headline or u.resume_headline or '',
        u.applications.count(),
        u.created_at.strftime('%Y-%m-%d') if u.created_at else '',
    ] for u in applicants]

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Applicants'
        header_fill = PatternFill('solid', fgColor='198754')
        header_font = Font(bold=True, color='FFFFFF')
        for col, h in enumerate(col_headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="applicants.xlsx"'}
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(col_headers)
    for row in rows:
        writer.writerow(row)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename="applicants.csv"'}
    )


# ─── SHARED: Company job alert helper ────────────────────────────────────────

def _send_company_job_alerts(position):
    """Email & notify about a new position.
    - Internships: university coordinators + student followers.
    - Regular jobs: all company followers.
    """
    from helpers import send_email
    from flask import current_app, url_for
    from flask import render_template as _rt
    if not position.company_id:
        return
    if position.type == 'Internship':
        site_url = current_app.config.get('SITE_URL', '')
        notified_ids = set()
        coord_memberships = UniversityMember.query.filter_by(role='coordinator').all()
        coordinators = User.query.filter(
            User.id.in_([m.user_id for m in coord_memberships]),
            User.is_active == True
        ).all() if coord_memberships else []
        for coord in coordinators:
            push_notification(coord.id,
                f'New internship posted: {position.title} at {position.company.name}',
                url_for('jobs.detail', job_id=position.id), 'bi-mortarboard-fill')
            try:
                html = _rt('emails/new_internship_alert.html',
                    recipient=coord, position=position,
                    company=position.company, is_coordinator=True, site_url=site_url)
                send_email(coord.email,
                    f'New internship: {position.title} at {position.company.name}', html)
            except Exception as e:
                current_app.logger.warning(f'Internship coordinator alert failed for {coord.email}: {e}')
            notified_ids.add(coord.id)
        student_follows = (
            CompanyFollow.query
            .join(User, User.id == CompanyFollow.user_id)
            .filter(CompanyFollow.company_id == position.company_id,
                    User.role == ROLE_STUDENT, User.is_active == True).all()
        )
        for follow in student_follows:
            student = follow.user
            if student.id in notified_ids:
                continue
            push_notification(student.id,
                f'New internship at {position.company.name}: {position.title}',
                url_for('jobs.detail', job_id=position.id), 'bi-mortarboard-fill')
            try:
                html = _rt('emails/new_internship_alert.html',
                    recipient=student, position=position,
                    company=position.company, is_coordinator=False, site_url=site_url)
                send_email(student.email,
                    f'New internship at {position.company.name}: {position.title}', html)
            except Exception as e:
                current_app.logger.warning(f'Internship student alert failed for {student.email}: {e}')
            notified_ids.add(student.id)
        if notified_ids:
            db.session.commit()
        return
    followers = CompanyFollow.query.filter_by(company_id=position.company_id).all()
    for follow in followers:
        user = follow.user
        try:
            html = _rt(
                'emails/new_job_alert.html',
                user=user, position=position, company=position.company,
            )
            send_email(
                user.email,
                f'New job at {position.company.name}: {position.title}',
                html,
            )
        except Exception as e:
            current_app.logger.warning(
                f'Job alert email failed for {user.email}: {e}')
        push_notification(
            user.id,
            f'New job at {position.company.name}: {position.title}',
            url_for('jobs.detail', job_id=position.id),
            'bi-briefcase-fill',
        )
    if followers:
        db.session.commit()