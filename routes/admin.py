import csv
import io
import secrets
from datetime import datetime, timedelta
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, abort, current_app, send_from_directory,
                   Response, send_file, jsonify)
from flask_login import current_user

from models import (db, User, Position, Application, ApplicationHistory,
                    Interview, AuditLog, Company, CompanyMember, CompanyFollow,
                    UserSkill, UserExperience, UserEducation,
                    UserLanguage, UserCertification,
                    Message, SupervisorRequest, University, UniversityDepartment, UniversityMember, UniversityRequest,
                    ROLE_ADMIN, ROLE_SUPERVISOR, ROLE_USER, ROLE_STUDENT, ROLE_UNIVERSITY_COORD,
                    LANG_LEVELS, ALL_STATUSES, SOURCES, STATUS_NEW, STATUS_UNIV_PENDING)
from sqlalchemy import or_, and_
from helpers import (admin_required, log_history, save_cv, allowed_file,
                     send_email, save_company_image, push_notification, log_audit,
                     get_site_settings, save_site_settings)


def _audit(action, target=''):
    """Write a system audit log entry (requires authenticated admin context)."""
    log_audit(action, target)


def _parse_date(s):
    """Return a date object from an ISO string, or None."""
    from datetime import date as _d
    try:
        return _d.fromisoformat(s.strip()) if s and s.strip() else None
    except ValueError:
        return None


def _parse_int(s):
    """Return an int from a string, or None."""
    try:
        return int(str(s).strip()) if s and str(s).strip() else None
    except (ValueError, TypeError):
        return None


def _delete_user_with_related_data(user):
    """Delete a user and related records used by admin delete flows."""
    for app in user.applications.all():
        app.history.order_by(None).delete(synchronize_session=False)
        app.interviews.delete(synchronize_session=False)
        db.session.delete(app)
    Message.query.filter(
        (Message.sender_id == user.id) | (Message.receiver_id == user.id)
    ).delete(synchronize_session=False)
    db.session.delete(user)


admin_bp = Blueprint('admin', __name__)


_ADMIN_UNIVERSITY_STUDENT_COLS = [
    ('student_id_number', 'Student ID'),
    ('full_name', 'Full Name'),
    ('email', 'Email'),
    ('phone', 'Phone'),
    ('department_name', 'Department'),
    ('university_class', 'Class'),
    ('university_major', 'Major'),
    ('graduation_year', 'Graduation Year'),
]


def _normalize_text(value):
    return ' '.join(str(value or '').strip().lower().split())


def _admin_university_students_query(univ_id, search='', department_id=None, class_filter='', major_filter='', graduation_year=None):
    q = User.query.filter_by(university_id=univ_id, role=ROLE_STUDENT)

    if search:
        q = q.filter(or_(
            User.full_name.ilike(f'%{search}%'),
            User.email.ilike(f'%{search}%'),
            User.university_major.ilike(f'%{search}%'),
            User.university_class.ilike(f'%{search}%'),
            User.student_id_number.ilike(f'%{search}%'),
        ))
    if department_id:
        q = q.filter(User.university_department_id == department_id)
    if class_filter:
        q = q.filter(User.university_class.ilike(f'%{class_filter}%'))
    if major_filter:
        q = q.filter(User.university_major.ilike(f'%{major_filter}%'))
    if graduation_year:
        q = q.filter(User.graduation_year == graduation_year)

    return q


def _admin_university_student_redirect(univ_id):
    params = {}
    for key in ('tab', 'q', 'department_id', 'class_filter', 'major_filter', 'graduation_year', 'page'):
        value = request.values.get(key, '').strip() if key != 'page' else request.values.get(key, type=int)
        if value:
            params[key] = value
    return redirect(url_for('admin.university_detail', univ_id=univ_id, **params))


def _resolve_university_department(univ_id, department_id=None, department_name=None):
    if department_id:
        return UniversityDepartment.query.filter_by(id=department_id, university_id=univ_id).first()
    if department_name:
        departments = UniversityDepartment.query.filter_by(university_id=univ_id).all()
        wanted = _normalize_text(department_name)
        for department in departments:
            if _normalize_text(department.name) == wanted or _normalize_text(department.full_name()) == wanted:
                return department
    return None


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@admin_bp.route('/')
@admin_required
def dashboard():
    from datetime import timedelta
    total_apps      = Application.query.count()
    new_last_7      = Application.query.filter(
        Application.applied_at >= datetime.utcnow() - timedelta(days=7)
    ).count()
    interview_count = Application.query.filter_by(status='Interview').count()
    hired_count     = Application.query.filter_by(status='Hired').count()

    # Status breakdown
    status_counts = {}
    for s in ALL_STATUSES:
        status_counts[s] = Application.query.filter_by(status=s).count()

    # Source breakdown
    source_counts = {}
    for src in SOURCES:
        source_counts[src] = Application.query.filter_by(source=src).count()

    # Latest 10 applications
    latest = (Application.query
              .order_by(Application.applied_at.desc())
              .limit(10).all())

    # Companies & supervisors for new stats
    companies_count   = Company.query.filter_by(is_active=True).count()
    supervisors_count = (User.query
                         .filter_by(role=ROLE_SUPERVISOR, is_active=True)
                         .count())
    total_positions   = Position.query.filter_by(is_active=True).count()
    total_users       = User.query.filter_by(role=ROLE_USER, is_active=True).count()

    # University-related metrics
    universities_count  = University.query.filter_by(is_active=True).count()
    students_count      = User.query.filter_by(role=ROLE_STUDENT, is_active=True).count()
    coordinators_count  = User.query.filter_by(role=ROLE_UNIVERSITY_COORD, is_active=True).count()
    internships_count   = (Application.query
                           .join(Position, Application.position_id == Position.id)
                           .filter(Position.type == 'Internship').count())
    univ_pending_count  = Application.query.filter_by(status=STATUS_UNIV_PENDING).count()

    recent_companies  = (Company.query
                         .filter_by(is_active=True)
                         .order_by(Company.id.desc())
                         .limit(5).all())
    recent_supervisors = (User.query
                          .filter_by(role=ROLE_SUPERVISOR, is_active=True)
                          .order_by(User.id.desc())
                          .limit(5).all())

    return render_template('admin/dashboard.html',
        total_apps=total_apps, new_last_7=new_last_7,
        interview_count=interview_count, hired_count=hired_count,
        status_counts=status_counts, source_counts=source_counts,
        latest=latest,
        companies_count=companies_count,
        supervisors_count=supervisors_count,
        total_positions=total_positions,
        total_users=total_users,
        universities_count=universities_count,
        students_count=students_count,
        coordinators_count=coordinators_count,
        internships_count=internships_count,
        univ_pending_count=univ_pending_count,
        recent_companies=recent_companies,
        recent_supervisors=recent_supervisors)


# ─── POSITIONS ────────────────────────────────────────────────────────────────

POSITION_TYPES = ['Full-time', 'Part-time', 'Contract', 'Internship', 'Remote', 'Freelance']


def _normalize_scope(scope):
    scope = (scope or 'all').strip().lower()
    return scope if scope in ('all', 'job', 'internship') else 'all'


def _apply_position_scope(query, scope):
    if scope == 'internship':
        return query.filter(Position.type == 'Internship')
    if scope == 'job':
        return query.filter(Position.type != 'Internship')
    return query


def _scope_meta(scope, kind='positions'):
    if kind == 'positions':
        if scope == 'job':
            return {
                'title': 'Job Positions',
                'item_singular': 'job position',
                'item_plural': 'job positions',
                'base_endpoint': 'admin.positions_jobs',
            }
        if scope == 'internship':
            return {
                'title': 'Internship Positions',
                'item_singular': 'internship position',
                'item_plural': 'internship positions',
                'base_endpoint': 'admin.positions_internships',
            }
        return {
            'title': 'Positions',
            'item_singular': 'position',
            'item_plural': 'positions',
            'base_endpoint': 'admin.positions',
        }

    if scope == 'job':
        return {
            'title': 'Job Applications',
            'item_singular': 'job application',
            'item_plural': 'job applications',
            'base_endpoint': 'admin.applications_jobs',
        }
    if scope == 'internship':
        return {
            'title': 'Internship Applications',
            'item_singular': 'internship application',
            'item_plural': 'internship applications',
            'base_endpoint': 'admin.applications_internships',
        }
    return {
        'title': 'Applications',
        'item_singular': 'application',
        'item_plural': 'applications',
        'base_endpoint': 'admin.applications',
    }


def _positions_list(scope='all'):
    scope = _normalize_scope(scope)
    page       = request.args.get('page', 1, type=int)
    q_str      = request.args.get('q', '').strip()
    status_f   = request.args.get('status', '')
    types_f    = request.args.getlist('type')
    company_f  = request.args.get('company_id', 0, type=int)
    dept_f     = request.args.get('dept', '').strip()

    q = _apply_position_scope(Position.query, scope)
    if q_str:
        q = q.filter(or_(Position.title.ilike(f'%{q_str}%'),
                         Position.department.ilike(f'%{q_str}%')))
    if status_f == 'active':
        q = q.filter_by(is_active=True)
    elif status_f == 'inactive':
        q = q.filter_by(is_active=False)
    if types_f:
        q = q.filter(Position.type.in_(types_f))
    if company_f:
        q = q.filter_by(company_id=company_f)
    if dept_f:
        q = q.filter(Position.department.ilike(f'%{dept_f}%'))

    positions = q.order_by(Position.created_at.desc()).paginate(
        page=page, per_page=current_app.config.get('POSITIONS_PER_PAGE', 25))

    all_companies = Company.query.filter_by(is_active=True).order_by(Company.name).all()
    scope_q = _apply_position_scope(Position.query, scope)

    kpi_active    = scope_q.filter_by(is_active=True).count()
    kpi_inactive  = scope_q.filter_by(is_active=False).count()
    if scope == 'internship':
        kpi_total_apps = (Application.query
                          .join(Position, Application.position_id == Position.id)
                          .filter(Position.type == 'Internship')
                          .count())
    elif scope == 'job':
        kpi_total_apps = (Application.query
                          .join(Position, Application.position_id == Position.id)
                          .filter(Position.type != 'Internship')
                          .count())
    else:
        kpi_total_apps = Application.query.count()
    kpi_closing   = scope_q.filter(
        Position.is_active == True,
        Position.closes_at != None,
        Position.closes_at <= datetime.utcnow() + timedelta(days=14)
    ).count()
    kpi_new_30d   = scope_q.filter(
        Position.created_at >= datetime.utcnow() - timedelta(days=30)
    ).count()

    meta = _scope_meta(scope, kind='positions')
    return render_template(
        'admin/positions.html',
        positions=positions,
        q_str=q_str,
        status_f=status_f,
        types_f=types_f,
        company_f=company_f,
        dept_f=dept_f,
        all_companies=all_companies,
        POSITION_TYPES=POSITION_TYPES,
        kpi_active=kpi_active,
        kpi_inactive=kpi_inactive,
        kpi_total_apps=kpi_total_apps,
        kpi_closing=kpi_closing,
        kpi_new_30d=kpi_new_30d,
        scope=scope,
        scope_title=meta['title'],
        item_singular=meta['item_singular'],
        item_plural=meta['item_plural'],
        positions_base_endpoint=meta['base_endpoint'],
    )


def _applications_list(scope='all'):
    scope = _normalize_scope(scope)
    page       = request.args.get('page', 1, type=int)
    status_f   = request.args.getlist('status')
    position_f = request.args.get('position_id', 0, type=int)
    source_f   = request.args.getlist('source')
    assigned_f = request.args.get('assigned_to', 0, type=int)
    search     = request.args.get('q', '').strip()

    q = (Application.query
         .join(User, Application.applicant_id == User.id)
         .join(Position, Application.position_id == Position.id))
    q = _apply_position_scope(q, scope)

    if status_f:
        q = q.filter(Application.status.in_(status_f))
    if position_f:
        q = q.filter(Application.position_id == position_f)
    if source_f:
        q = q.filter(Application.source.in_(source_f))
    if assigned_f:
        q = q.filter(Application.assigned_to_id == assigned_f)
    if search:
        q = q.filter(or_(User.full_name.ilike(f'%{search}%'),
                         User.email.ilike(f'%{search}%')))

    apps = q.order_by(Application.applied_at.desc()).paginate(
        page=page, per_page=current_app.config.get('APPLICATIONS_PER_PAGE', 25))

    all_positions = _apply_position_scope(Position.query, scope).order_by(Position.title).all()
    supervisors = User.query.filter_by(role=ROLE_SUPERVISOR, is_active=True).order_by(User.full_name).all()

    scope_apps = _apply_position_scope(
        Application.query.join(Position, Application.position_id == Position.id), scope
    )
    kpi_total      = scope_apps.count()
    kpi_new_7d     = scope_apps.filter(
        Application.applied_at >= datetime.utcnow() - timedelta(days=7)).count()
    kpi_interview  = scope_apps.filter(Application.status == 'Interview').count()
    kpi_hired      = scope_apps.filter(Application.status == 'Hired').count()
    kpi_rejected   = scope_apps.filter(Application.status == 'Rejected').count()
    kpi_pending    = scope_apps.filter(Application.status == 'New').count()

    meta = _scope_meta(scope, kind='applications')
    return render_template(
        'admin/applications.html',
        apps=apps,
        all_positions=all_positions,
        supervisors=supervisors,
        ALL_STATUSES=ALL_STATUSES,
        SOURCES=SOURCES,
        status_f=status_f,
        position_f=position_f,
        source_f=source_f,
        assigned_f=assigned_f,
        search=search,
        kpi_total=kpi_total,
        kpi_new_7d=kpi_new_7d,
        kpi_interview=kpi_interview,
        kpi_hired=kpi_hired,
        kpi_rejected=kpi_rejected,
        kpi_pending=kpi_pending,
        scope=scope,
        scope_title=meta['title'],
        item_singular=meta['item_singular'],
        item_plural=meta['item_plural'],
        applications_base_endpoint=meta['base_endpoint'],
    )

@admin_bp.route('/positions')
@admin_required
def positions():
    return _positions_list(scope='all')


@admin_bp.route('/positions/jobs')
@admin_required
def positions_jobs():
    return _positions_list(scope='job')


@admin_bp.route('/positions/internships')
@admin_required
def positions_internships():
    return _positions_list(scope='internship')


@admin_bp.route('/positions/<int:pos_id>')
@admin_required
def position_detail(pos_id):
    pos = db.get_or_404(Position, pos_id)
    recent_apps = pos.applications.order_by(Application.applied_at.desc()).limit(5).all()
    scope = _normalize_scope(request.args.get('scope', 'all'))
    back_endpoint = _scope_meta(scope, kind='positions')['base_endpoint']
    return render_template('admin/position_detail.html', pos=pos, recent_apps=recent_apps,
                           scope=scope, back_endpoint=back_endpoint)


@admin_bp.route('/positions/export')
@admin_required
def positions_export():
    fmt = request.args.get('fmt', 'csv')
    scope = _normalize_scope(request.args.get('scope', 'all'))
    rows = _apply_position_scope(Position.query, scope).order_by(Position.created_at.desc()).all()
    headers = ['ID', 'Title', 'Department', 'Type', 'Location', 'Company',
               'Status', 'Applications', 'Views', 'Created', 'Closes']
    data = [[
        p.id, p.title, p.department or '', p.type, p.location or '',
        p.company.name if p.company else '',
        'Active' if p.is_active else 'Inactive',
        p.application_count, p.views_count,
        p.created_at.strftime('%Y-%m-%d'),
        p.closes_at.strftime('%Y-%m-%d') if p.closes_at else ''
    ] for p in rows]
    filename = 'positions'
    if scope == 'job':
        filename = 'job_positions'
    elif scope == 'internship':
        filename = 'internship_positions'
    return _export(headers, data, filename, fmt)


@admin_bp.route('/positions/new', methods=['GET', 'POST'])
@admin_required
def position_new():
    if request.method == 'POST':
        company_id = request.form.get('company_id', type=int) or None
        pos = Position(
            title        = request.form['title'].strip(),
            department   = request.form.get('department', '').strip(),
            location     = request.form.get('location', 'Slemani, Iraq').strip(),
            type         = request.form.get('type', 'Full-time'),
            description  = request.form.get('description', '').strip(),
            requirements = request.form.get('requirements', '').strip(),
            salary_range = request.form.get('salary_range', '').strip(),
            is_active    = bool(request.form.get('is_active')),
            company_id   = company_id,
            created_by   = current_user.id,
        )
        closes = request.form.get('closes_at', '').strip()
        if closes:
            pos.closes_at = datetime.strptime(closes, '%Y-%m-%d')
        db.session.add(pos)
        _audit('position.create', pos.title)
        db.session.commit()
        notify = request.form.get('notify_users') == '1'
        if pos.is_active and company_id and notify:
            _send_company_job_alerts(pos)
        flash(f'Position "{pos.title}" created.', 'success')
        return redirect(url_for('admin.positions'))
    companies = Company.query.filter_by(is_active=True).order_by(Company.name).all()
    return render_template('admin/position_form.html', pos=None, companies=companies)


@admin_bp.route('/positions/<int:pos_id>/edit', methods=['GET', 'POST'])
@admin_required
def position_edit(pos_id):
    pos = db.get_or_404(Position, pos_id)
    if request.method == 'POST':
        was_active   = pos.is_active
        old_company  = pos.company_id
        company_id   = request.form.get('company_id', type=int) or None
        pos.title        = request.form['title'].strip()
        pos.department   = request.form.get('department', '').strip()
        pos.location     = request.form.get('location', '').strip()
        pos.type         = request.form.get('type', 'Full-time')
        pos.description  = request.form.get('description', '').strip()
        pos.requirements = request.form.get('requirements', '').strip()
        pos.salary_range = request.form.get('salary_range', '').strip()
        pos.is_active    = bool(request.form.get('is_active'))
        pos.company_id   = company_id
        closes = request.form.get('closes_at', '').strip()
        pos.closes_at    = datetime.strptime(closes, '%Y-%m-%d') if closes else None
        _audit('position.edit', pos.title)
        db.session.commit()
        # Send alerts only if admin explicitly checks notify
        notify = request.form.get('notify_users') == '1'
        if pos.is_active and company_id and notify:
            _send_company_job_alerts(pos)
        flash('Position updated.', 'success')
        return redirect(url_for('admin.positions'))
    companies = Company.query.filter_by(is_active=True).order_by(Company.name).all()
    return render_template('admin/position_form.html', pos=pos, companies=companies)


@admin_bp.route('/positions/<int:pos_id>/toggle', methods=['POST'])
@admin_required
def position_toggle(pos_id):
    pos = db.get_or_404(Position, pos_id)
    pos.is_active = not pos.is_active
    db.session.commit()
    state = 'activated' if pos.is_active else 'deactivated'
    flash(f'"{pos.title}" {state}.', 'success')
    return redirect(request.referrer or url_for('admin.positions'))


@admin_bp.route('/positions/<int:pos_id>/delete', methods=['POST'])
@admin_required
def position_delete(pos_id):
    pos = db.get_or_404(Position, pos_id)
    title = pos.title
    app_count = pos.application_count
    for app in pos.applications.all():
        app.history.delete(synchronize_session=False)
        app.interviews.delete(synchronize_session=False)
        db.session.delete(app)
    db.session.delete(pos)
    _audit('position.delete', f'{title} ({app_count} apps deleted)')
    db.session.commit()
    flash(f'Position "{title}" and {app_count} application(s) permanently deleted.', 'success')
    return redirect(url_for('admin.positions'))


# ─── APPLICATIONS ─────────────────────────────────────────────────────────────

@admin_bp.route('/applications')
@admin_required
def applications():
    return _applications_list(scope='all')


@admin_bp.route('/applications/jobs')
@admin_required
def applications_jobs():
    return _applications_list(scope='job')


@admin_bp.route('/applications/internships')
@admin_required
def applications_internships():
    return _applications_list(scope='internship')


@admin_bp.route('/applications/export')
@admin_required
def applications_export():
    from models import UserSkill, UserExperience, UserEducation
    fmt = request.args.get('fmt', 'csv')
    scope = _normalize_scope(request.args.get('scope', 'all'))
    rows = (_apply_position_scope(
        Application.query.join(Position, Application.position_id == Position.id), scope
    ).order_by(Application.applied_at.desc()).all())

    headers = [
        # ── Application submission info ──────────────────────────────────
        'App ID', 'Applied At', 'Updated At', 'Status', 'Source',
        'Position', 'Department', 'Position Type', 'Company',
        'Expected Salary',
        'Assigned To',
        'Cover Letter', 'CV Filename (Original)',
        'Internal Notes',
        # ── Applicant identity ───────────────────────────────────────────
        'Full Name', 'Email', 'Phone',
        'Headline', 'Resume Summary',
        'Location', 'Nationality', 'Gender', 'Date of Birth',
        # ── Online presence ──────────────────────────────────────────────
        'LinkedIn', 'Portfolio', 'GitHub',
        # ── Profile sections ─────────────────────────────────────────────
        'Skills',
        'Latest Experience (Title @ Company)',
        'All Experience',
        'Education',
        'Languages',
        'Certifications',
        # ── Misc ─────────────────────────────────────────────────────────
        'Profile Strength (%)', 'Bio',
    ]

    data = []
    for a in rows:
        u = a.applicant

        # Skills
        skills_str = ', '.join(
            f'{s.name} ({s.proficiency})' for s in
            UserSkill.query.filter_by(user_id=u.id).order_by(UserSkill.name).all()
        )

        # Experiences
        exps = UserExperience.query.filter_by(user_id=u.id).order_by(
            UserExperience.start_date.desc().nullslast()).all()
        latest_exp = ''
        if exps:
            e = exps[0]
            latest_exp = f'{e.title}' + (f' @ {e.company}' if e.company else '')
        all_exps = ' | '.join(
            f'{e.title}' + (f' @ {e.company}' if e.company else '') +
            (f' ({e.start_date.year}–{e.end_date.year if e.end_date else "present"})' if e.start_date else '')
            for e in exps
        )

        # Education
        edus = UserEducation.query.filter_by(user_id=u.id).order_by(
            UserEducation.end_year.desc().nullslast()).all()
        edu_str = ' | '.join(
            f'{e.degree or ""}{"," if e.degree and e.field else ""}{e.field or ""} — {e.institution}'
            + (f' ({e.end_year})' if e.end_year else '')
            for e in edus
        )

        # Languages removed (rollback)
        langs_str = ''

        # Certifications removed (rollback)
        certs_str = ''

        dob = u.date_of_birth.strftime('%Y-%m-%d') if u.date_of_birth else ''

        data.append([
            a.id,
            a.applied_at.strftime('%Y-%m-%d %H:%M'),
            a.updated_at.strftime('%Y-%m-%d %H:%M') if a.updated_at else '',
            a.status,
            a.source,
            a.position.title,
            a.position.department or '',
            a.position.type,
            a.position.company.name if a.position.company else '',
            a.expected_salary or '',
            a.assigned_to.full_name if a.assigned_to else '',
            a.cover_letter or '',
            a.cv_original or '',
            a.notes or '',
            u.full_name,
            u.email,
            u.phone or '',
            u.headline or '',
            u.resume_headline or '',
            u.location_city or '',
            u.nationality or '',
            u.gender or '',
            dob,
            u.linkedin_url or '',
            u.portfolio_url or '',
            u.github_url or '',
            skills_str,
            latest_exp,
            all_exps,
            edu_str,
            langs_str,
            certs_str,
            u.profile_strength,
            (u.bio or '').replace('\n', ' '),
        ])

    filename = 'applications'
    if scope == 'job':
        filename = 'job_applications'
    elif scope == 'internship':
        filename = 'internship_applications'
    return _export(headers, data, filename, fmt)


@admin_bp.route('/applications/<int:app_id>')
@admin_required
def application_detail(app_id):
    app = db.get_or_404(Application, app_id)
    scope = _normalize_scope(request.args.get('scope', 'all'))
    back_endpoint = _scope_meta(scope, kind='applications')['base_endpoint']
    is_internship = bool(app.position and app.position.type == 'Internship')
    assignee_role = ROLE_UNIVERSITY_COORD if is_internship else ROLE_SUPERVISOR
    assignee_label = 'Coordinator' if is_internship else 'Supervisor'
    if is_internship:
        # Only allow coordinators that belong to the applicant's university.
        applicant_univ_id = getattr(app.applicant, 'university_id', None)
        if applicant_univ_id:
            coord_ids = [m.user_id for m in UniversityMember.query
                         .filter_by(university_id=applicant_univ_id, role='coordinator').all()]
            if coord_ids:
                supervisors = (User.query
                               .filter(User.id.in_(coord_ids),
                                       User.role == ROLE_UNIVERSITY_COORD,
                                       User.is_active == True)
                               .order_by(User.full_name).all())
            else:
                supervisors = []
        else:
            # Student not linked to any university — no valid coordinator can be assigned.
            supervisors = []
    else:
        supervisors = User.query.filter_by(role=assignee_role, is_active=True).all()
    # Admins see all status changes + admin-authored entries (notes without status change)
    admin_ids = [u.id for u in User.query.filter_by(role=ROLE_ADMIN).with_entities(User.id).all()]
    history   = app.history.filter(
        db.or_(
            ApplicationHistory.new_status.isnot(None),
            ApplicationHistory.changed_by_id.in_(admin_ids)
        )
    ).order_by(ApplicationHistory.created_at.desc()).all()
    interviews  = app.interviews.order_by(Interview.scheduled_at).all()

    # Applicant profile
    u = app.applicant
    skills         = UserSkill.query.filter_by(user_id=u.id).order_by(UserSkill.name).all()
    experiences    = UserExperience.query.filter_by(user_id=u.id).order_by(
                        UserExperience.start_date.desc().nullslast()).all()
    educations     = UserEducation.query.filter_by(user_id=u.id).order_by(
                        UserEducation.end_year.desc().nullslast()).all()
# Removed languages and certifications (rollback)
    languages = []
    certifications = []
    # Message thread between current admin and the applicant
    thread_messages = (Message.query
                       .filter(or_(
                           and_(Message.sender_id == current_user.id,
                                Message.receiver_id == u.id),
                           and_(Message.sender_id == u.id,
                                Message.receiver_id == current_user.id)
                       ))
                       .order_by(Message.created_at.asc())
                       .all())
    # Mark received as read
    Message.query.filter_by(sender_id=u.id, receiver_id=current_user.id,
                            is_read=False).update({'is_read': True})

    # Message thread between ANY admin and the assigned supervisor (if any)
    supervisor_thread = []
    if app.assigned_to_id:
        admin_ids = [u.id for u in User.query.filter_by(role=ROLE_ADMIN, is_active=True).all()]
        supervisor_thread = (Message.query
                             .filter(or_(
                                 and_(Message.sender_id == app.assigned_to_id,
                                      Message.receiver_id.in_(admin_ids)),
                                 and_(Message.sender_id.in_(admin_ids),
                                      Message.receiver_id == app.assigned_to_id)
                             ))
                             .order_by(Message.created_at.asc())
                             .all())
        Message.query.filter(
            Message.sender_id == app.assigned_to_id,
            Message.receiver_id.in_(admin_ids),
            Message.is_read == False
        ).update({'is_read': True}, synchronize_session=False)

    db.session.commit()

    available_statuses = ALL_STATUSES if (app.position and app.position.type == 'Internship') else [s for s in ALL_STATUSES if s != STATUS_UNIV_PENDING]
    return render_template('admin/application_detail.html',
        app=app, supervisors=supervisors, ALL_STATUSES=available_statuses,
        history=history, interviews=interviews,
        skills=skills, experiences=experiences, educations=educations,
        languages=languages, certifications=certifications,
        thread_messages=thread_messages, supervisor_thread=supervisor_thread,
        scope=scope, back_endpoint=back_endpoint,
        is_internship=is_internship, assignee_label=assignee_label)


@admin_bp.route('/applications/<int:app_id>/update', methods=['POST'])
@admin_required
def application_update(app_id):
    application = db.get_or_404(Application, app_id)
    new_status  = request.form.get('status')
    note        = request.form.get('note', '').strip()
    assign_to   = request.form.get('assigned_to_id', type=int)

    status_changed = new_status and new_status != application.status

    if status_changed:
        log_history(application, current_user, new_status=new_status, note=note or None, is_internal=True)
        application.status = new_status
        application.updated_at = datetime.utcnow()
        _audit('application.status_change', f'#{application.id} → {new_status}')
    elif note:
        log_history(application, current_user, note=note, is_internal=True)

    if assign_to is not None:
        old_assignee = application.assigned_to_id
        # Safety: for internships, only allow coordinators that belong to the
        # applicant's university. Prevents misassignment via crafted form posts.
        if assign_to:
            is_internship = bool(application.position and application.position.type == 'Internship')
            if is_internship:
                applicant_univ_id = getattr(application.applicant, 'university_id', None)
                if not applicant_univ_id:
                    flash('Cannot assign a coordinator: this student is not linked to any university.', 'danger')
                    return redirect(url_for('admin.application_detail', app_id=app_id))
                allowed = UniversityMember.query.filter_by(
                    university_id=applicant_univ_id, user_id=assign_to, role='coordinator').first()
                if not allowed:
                    flash('That coordinator does not belong to the student\'s university and cannot be assigned.', 'danger')
                    return redirect(url_for('admin.application_detail', app_id=app_id))
        application.assigned_to_id = assign_to or None
        if assign_to and assign_to != old_assignee:
            supervisor = db.session.get(User, assign_to)
            if supervisor:
                _audit('application.assign', f'#{application.id} → {supervisor.full_name}')
                _send_assignment_email(application, supervisor)
                push_notification(
                    supervisor.id,
                    f'Application #{application.id} ({application.position.title}) assigned to you',
                    url_for('supervisor.application_detail', app_id=application.id)
                )

    db.session.commit()

    # Send email + push notification AFTER commit so new status is fully persisted
    if status_changed:
        _send_status_email(application)
        push_notification(
            application.applicant_id,
            f'Your application for "{application.position.title}" is now: {new_status}',
            url_for('user.my_applications')
        )

    flash('Application updated.', 'success')
    return redirect(url_for('admin.application_detail', app_id=app_id))


@admin_bp.route('/applications/<int:app_id>/delete', methods=['POST'])
@admin_required
def application_delete(app_id):
    application = db.get_or_404(Application, app_id)
    application.history.order_by(None).delete(synchronize_session=False)
    application.interviews.delete(synchronize_session=False)
    db.session.delete(application)
    _audit('application.delete', f'#{app_id} {application.position.title}')
    db.session.commit()
    flash('Application permanently deleted.', 'success')
    return redirect(request.referrer or url_for('admin.applications'))


@admin_bp.route('/applications/<int:app_id>/history/<int:history_id>/delete', methods=['POST'])
@admin_required
def history_delete(app_id, history_id):
    entry = db.get_or_404(ApplicationHistory, history_id)
    if entry.application_id != app_id:
        abort(404)
    db.session.delete(entry)
    _audit('application.history_delete', f'app #{app_id} history #{history_id}')
    db.session.commit()
    flash('Timeline entry deleted.', 'success')
    return redirect(url_for('admin.application_detail', app_id=app_id))


@admin_bp.route('/applications/<int:app_id>/interview', methods=['POST'])
@admin_required
def schedule_interview(app_id):
    application = db.get_or_404(Application, app_id)
    scheduled   = request.form.get('scheduled_at', '')
    location    = request.form.get('location', '').strip()
    notes       = request.form.get('notes', '').strip()

    if not scheduled:
        flash('Please provide a date/time for the interview.', 'danger')
        return redirect(url_for('admin.application_detail', app_id=app_id))

    interview = Interview(
        application_id  = application.id,
        scheduled_at    = datetime.strptime(scheduled, '%Y-%m-%dT%H:%M'),
        location        = location,
        interviewer_id  = current_user.id,
        notes           = notes,
    )
    db.session.add(interview)

    if application.status != 'Interview':
        log_history(application, current_user, new_status='Interview',
                    note=f'Interview scheduled for {scheduled}')
        application.status = 'Interview'

    db.session.commit()
    flash('Interview scheduled.', 'success')
    return redirect(url_for('admin.application_detail', app_id=app_id))


# ─── USERS ────────────────────────────────────────────────────────────────────

@admin_bp.route('/users')
@admin_required
def users():
    page     = request.args.get('page', 1, type=int)
    q_str    = request.args.get('q', '').strip()
    roles_f  = request.args.getlist('role')
    status_f = request.args.get('status', '')

    q = User.query
    if q_str:
        q = q.filter(or_(User.full_name.ilike(f'%{q_str}%'),
                         User.email.ilike(f'%{q_str}%')))
    if roles_f:
        q = q.filter(User.role.in_(roles_f))
    if status_f == 'active':
        q = q.filter_by(is_active=True)
    elif status_f == 'inactive':
        q = q.filter_by(is_active=False)

    users_paged = q.order_by(User.role, User.full_name).paginate(
        page=page, per_page=30, error_out=False)

    kpi_total      = User.query.count()
    kpi_admins     = User.query.filter_by(role=ROLE_ADMIN).count()
    kpi_supervisors= User.query.filter_by(role=ROLE_SUPERVISOR).count()
    kpi_users      = User.query.filter_by(role=ROLE_USER).count()
    kpi_students   = User.query.filter_by(role=ROLE_STUDENT).count()
    kpi_coordinators = User.query.filter_by(role=ROLE_UNIVERSITY_COORD).count()
    kpi_active     = User.query.filter_by(is_active=True).count()
    kpi_inactive   = User.query.filter_by(is_active=False).count()

    return render_template('admin/users.html', users=users_paged,
                           q_str=q_str, roles_f=roles_f, status_f=status_f,
                           ROLES=[ROLE_ADMIN, ROLE_SUPERVISOR, ROLE_USER,
                                  ROLE_STUDENT, ROLE_UNIVERSITY_COORD],
                           kpi_total=kpi_total, kpi_admins=kpi_admins,
                           kpi_supervisors=kpi_supervisors, kpi_users=kpi_users,
                           kpi_students=kpi_students, kpi_coordinators=kpi_coordinators,
                           kpi_active=kpi_active, kpi_inactive=kpi_inactive)


@admin_bp.route('/users/export')
@admin_required
def users_export():
    fmt = request.args.get('fmt', 'csv')
    rows = User.query.order_by(User.role, User.full_name).all()
    headers = ['ID', 'Full Name', 'Email', 'Phone', 'Role', 'Status',
               'Created At', 'Last Login']
    data = [[
        u.id, u.full_name, u.email, u.phone or '',
        u.role, 'Active' if u.is_active else 'Inactive',
        u.created_at.strftime('%Y-%m-%d') if hasattr(u, 'created_at') and u.created_at else '',
        u.last_login.strftime('%Y-%m-%d %H:%M') if u.last_login else 'Never'
    ] for u in rows]
    return _export(headers, data, 'users', fmt)


@admin_bp.route('/users/<int:user_id>')
@admin_required
def user_detail(user_id):
    user = db.get_or_404(User, user_id)
    recent_apps = user.applications.order_by(Application.applied_at.desc()).limit(5).all()
    return render_template('admin/user_detail.html', user=user, recent_apps=recent_apps)


@admin_bp.route('/users/new', methods=['GET', 'POST'])
@admin_required
def user_new():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        if User.query.filter_by(email=email).first():
            flash('Email already exists.', 'danger')
            return redirect(url_for('admin.user_new'))

        user = User(
            full_name = request.form.get('full_name', '').strip(),
            email     = email,
            phone     = request.form.get('phone', '').strip(),
            role      = request.form.get('role', ROLE_USER),
            is_active = bool(request.form.get('is_active', True)),
            daily_new_apps_reminder = bool(request.form.get('daily_new_apps_reminder')),
            weekly_job_match_digest = bool(request.form.get('weekly_job_match_digest')),
            weekly_coord_digest     = bool(request.form.get('weekly_coord_digest')),
        )
        user.set_password(request.form.get('password', 'ChangeMe@123'))
        db.session.add(user)
        _audit('user.create', f'{user.full_name} <{user.email}>')
        db.session.commit()
        # Send welcome email based on role
        try:
            temp_pw = request.form.get('password', 'ChangeMe@123')
            if user.role in (ROLE_SUPERVISOR, ROLE_ADMIN):
                html = render_template('emails/welcome_supervisor.html',
                                       user=user,
                                       temp_password=temp_pw,
                                       is_admin=(user.role == ROLE_ADMIN))
                subject = 'Your MOF Jobs Admin Account is Ready' if user.role == ROLE_ADMIN else 'Your MOF Jobs Supervisor Account is Ready'
                send_email(user.email, subject, html)
            else:
                html = render_template('emails/welcome_user.html', user=user)
                send_email(user.email, 'Welcome to MOF Jobs — Your Account is Ready', html)
        except Exception as ex:
            current_app.logger.warning(f'User-create welcome email failed: {ex}')
        flash(f'User {user.full_name} created.', 'success')
        return redirect(url_for('admin.users'))
    return render_template('admin/user_form.html', user=None,
                           ROLES=[ROLE_ADMIN, ROLE_SUPERVISOR, ROLE_USER,
                                  ROLE_STUDENT, ROLE_UNIVERSITY_COORD])


@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def user_edit(user_id):
    user = db.get_or_404(User, user_id)
    if request.method == 'POST':
        # ── Account ──────────────────────────────────────────────────────────
        user.full_name = request.form.get('full_name', '').strip()
        new_email = request.form.get('email', '').strip().lower()
        if new_email and new_email != user.email:
            existing = User.query.filter(User.email == new_email, User.id != user.id).first()
            if existing:
                flash('That email is already in use by another account.', 'danger')
                return redirect(url_for('admin.user_edit', user_id=user.id))
            user.email = new_email
        user.phone     = request.form.get('phone', '').strip() or None
        user.role      = request.form.get('role', user.role)
        user.is_active = bool(request.form.get('is_active'))
        user.daily_new_apps_reminder = bool(request.form.get('daily_new_apps_reminder'))
        user.weekly_job_match_digest = bool(request.form.get('weekly_job_match_digest'))
        user.weekly_coord_digest     = bool(request.form.get('weekly_coord_digest'))
        new_pw = request.form.get('new_password', '').strip()
        if new_pw:
            user.set_password(new_pw)

        # ── Profile ───────────────────────────────────────────────────────────
        user.gender          = request.form.get('gender') or None
        user.location_city   = request.form.get('location_city', '').strip() or None
        user.headline        = request.form.get('headline', '').strip() or None
        user.bio             = request.form.get('bio', '').strip() or None
        user.resume_headline = request.form.get('resume_headline', '').strip() or None
        user.nationality     = request.form.get('nationality', '').strip() or None
        user.linkedin_url    = request.form.get('linkedin_url', '').strip() or None
        user.github_url      = request.form.get('github_url', '').strip() or None
        user.portfolio_url   = request.form.get('portfolio_url', '').strip() or None
        dob_str = request.form.get('date_of_birth', '').strip()
        user.date_of_birth = _parse_date(dob_str)

        # ── Skills ────────────────────────────────────────────────────────────
        UserSkill.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        for sname, sprof in zip(request.form.getlist('skill_name[]'),
                                request.form.getlist('skill_prof[]')):
            sname = sname.strip()
            if sname:
                db.session.add(UserSkill(user_id=user.id, name=sname,
                                         proficiency=sprof or 'intermediate'))

        # ── Work Experience ───────────────────────────────────────────────────
        UserExperience.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        for title, co, start, end, desc in zip(
                request.form.getlist('exp_title[]'),
                request.form.getlist('exp_company[]'),
                request.form.getlist('exp_start[]'),
                request.form.getlist('exp_end[]'),
                request.form.getlist('exp_desc[]')):
            title = title.strip()
            if title:
                db.session.add(UserExperience(
                    user_id=user.id, title=title,
                    company=co.strip() or None,
                    start_date=_parse_date(start),
                    end_date=_parse_date(end),
                    description=desc.strip() or None))

        # ── Education ────────────────────────────────────────────────────────
        UserEducation.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        for inst, deg, fld, sy, ey in zip(
                request.form.getlist('edu_inst[]'),
                request.form.getlist('edu_degree[]'),
                request.form.getlist('edu_field[]'),
                request.form.getlist('edu_start_year[]'),
                request.form.getlist('edu_end_year[]')):
            inst = inst.strip()
            if inst:
                db.session.add(UserEducation(
                    user_id=user.id, institution=inst,
                    degree=deg.strip() or None,
                    field=fld.strip() or None,
                    start_year=_parse_int(sy),
                    end_year=_parse_int(ey)))

        # ── Languages ────────────────────────────────────────────────────────
        UserLanguage.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        for lname, lprof in zip(request.form.getlist('lang_name[]'),
                                request.form.getlist('lang_prof[]')):
            lname = lname.strip()
            if lname:
                db.session.add(UserLanguage(user_id=user.id, language=lname,
                                             proficiency=lprof or 'Intermediate'))

        # ── Certifications ───────────────────────────────────────────────────────
        UserCertification.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        for cname, corg, cissued, ccid, curl in zip(
                request.form.getlist('cert_name[]'),
                request.form.getlist('cert_org[]'),
                request.form.getlist('cert_issued[]'),
                request.form.getlist('cert_cid[]'),
                request.form.getlist('cert_url[]')):
            cname = cname.strip()
            if cname:
                cert = UserCertification(
                    user_id=user.id, name=cname,
                    issuing_org=corg.strip() or None,
                    credential_id=ccid.strip() or None,
                    credential_url=curl.strip() or None,
                )
                if cissued:
                    try:
                        from datetime import datetime as _dt
                        cert.issue_date = _dt.strptime(cissued, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                db.session.add(cert)

        _audit('user.edit', f'{user.full_name} <{user.email}>')
        db.session.commit()
        flash('User updated.', 'success')
        return redirect(url_for('admin.users'))
    return render_template('admin/user_form.html', user=user,
                           ROLES=[ROLE_ADMIN, ROLE_SUPERVISOR, ROLE_USER,
                                  ROLE_STUDENT, ROLE_UNIVERSITY_COORD],
                           LANG_LEVELS=LANG_LEVELS)


@admin_bp.route('/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def user_toggle(user_id):
    user = db.get_or_404(User, user_id)
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'danger')
    else:
        user.is_active = not user.is_active
        db.session.commit()
        flash(f'{user.full_name} {"activated" if user.is_active else "deactivated"}.', 'success')
    return redirect(request.referrer or url_for('admin.users'))


@admin_bp.route('/users/<int:user_id>/send-reminder-test', methods=['POST'])
@admin_required
def user_send_reminder_test(user_id):
    """Send a one-off test of the supervisor daily reminder email.

    Sends the email directly to the SUPERVISOR (the actual receiver) so
    they get a real test of what they would normally receive each morning.
    If the supervisor has no "New" applications, a sample preview with
    mock data is sent so the layout is still visible.
    """
    from datetime import datetime, timedelta
    from sqlalchemy import or_
    from models import Application, Position, CompanyMember, STATUS_NEW
    user = db.get_or_404(User, user_id)

    # Build the same query the cron uses (same visibility rules)
    managed_company_ids = [m.company_id for m in
                            CompanyMember.query.filter_by(user_id=user.id, role='manager').all()]
    managed_pos_ids = []
    if managed_company_ids:
        managed_pos_ids = [p.id for p in
                            Position.query.filter(Position.company_id.in_(managed_company_ids)).all()]
    q = Application.query.filter(Application.status == STATUS_NEW)
    if managed_pos_ids:
        q = q.filter(or_(Application.position_id.in_(managed_pos_ids),
                         Application.assigned_to_id == user.id))
    else:
        q = q.filter(Application.assigned_to_id == user.id)
    apps = q.order_by(Application.applied_at.asc()).all()

    is_sample = False
    if not apps:
        # Build a sample list using ANY recent New applications so the admin
        # can still preview the email layout
        apps = (Application.query.filter(Application.status == STATUS_NEW)
                .order_by(Application.applied_at.desc()).limit(3).all())
        if not apps:
            apps = (Application.query
                    .order_by(Application.applied_at.desc()).limit(3).all())
        is_sample = True

    if not apps:
        return jsonify({'ok': False, 'error': 'No applications exist in the system at all — cannot generate a preview.'}), 400

    # Send to the actual supervisor (the receiver) so they get a real test of what they would normally receive
    target_email = (user.email or '').strip()
    if not target_email:
        return jsonify({'ok': False, 'error': 'This user has no email address on file.'}), 400
    try:
        now = datetime.utcnow()
        site_url = request.host_url.rstrip('/')
        html = render_template('emails/supervisor_daily_reminder.html',
                               supervisor=user, apps=apps, count=len(apps),
                               now=now, site_url=site_url)
        prefix = '[TEST · SAMPLE DATA] ' if is_sample else '[TEST] '
        subject = f'{prefix}Reminder: {len(apps)} applicant{"s" if len(apps) != 1 else ""} waiting for review'
        send_email(target_email, subject, html)
        return jsonify({
            'ok': True,
            'sent_to': target_email,
            'count': len(apps),
            'is_sample': is_sample,
            'message': f'Preview sent to {target_email}' + (' (using sample data)' if is_sample else ''),
        })
    except Exception as ex:
        current_app.logger.exception('Test reminder email failed: %s', ex)
        return jsonify({'ok': False, 'error': str(ex)}), 500


@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def user_delete(user_id):
    user = db.get_or_404(User, user_id)
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin.users'))
    name = user.full_name
    _delete_user_with_related_data(user)
    _audit('user.delete', f'{name}')
    db.session.commit()
    flash(f'User "{name}" permanently deleted.', 'success')
    return redirect(url_for('admin.users'))


# ─── CV DOWNLOAD ──────────────────────────────────────────────────────────────

@admin_bp.route('/cv/<filename>')
@admin_required
def download_cv(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _send_status_email(application):
    """Send applicant a status update email."""
    try:
        site_url = current_app.config.get('SITE_URL', 'https://jobs.mof-eng.com')
        html = render_template('emails/status_update.html', app=application, site_url=site_url)
        send_email(application.applicant.email,
                   f'Your application update — {application.position.title}',
                   html)
        current_app.logger.info(f'Status email sent for app #{application.id} to {application.applicant.email}')
    except Exception as e:
        import traceback
        current_app.logger.error(f'Status email FAILED for app #{application.id}: {e}\n{traceback.format_exc()}')
        flash(f'Status updated, but email to {application.applicant.email} failed: {e}', 'warning')


def _send_assignment_email(application, supervisor):
    """Notify supervisor that an application has been assigned to them."""
    try:
        html = render_template('emails/assignment_notification.html',
                               app=application, supervisor=supervisor)
        send_email(supervisor.email,
                   f'New application assigned to you — {application.position.title}',
                   html)
    except Exception as e:
        current_app.logger.warning(f'Assignment email failed: {e}')


# ─── AUDIT TRAIL ──────────────────────────────────────────────────────────────

@admin_bp.route('/audit')
@admin_required
def audit_trail():
    page = request.args.get('page', 1, type=int)
    user_f = request.args.get('user_id', 0, type=int)
    action_f = request.args.get('action', '')

    # System audit log
    q_audit = AuditLog.query.order_by(AuditLog.created_at.desc())
    if user_f:
        q_audit = q_audit.filter_by(user_id=user_f)
    if action_f:
        q_audit = q_audit.filter(AuditLog.action.ilike(f'%{action_f}%'))

    audit_entries = q_audit.paginate(page=page, per_page=40)
    all_users = User.query.order_by(User.full_name).all()

    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    kpi_total      = AuditLog.query.count()
    kpi_today      = AuditLog.query.filter(AuditLog.created_at >= today_start).count()
    kpi_this_week  = AuditLog.query.filter(
        AuditLog.created_at >= datetime.utcnow() - timedelta(days=7)).count()
    kpi_users_active = db.session.query(
        db.func.count(db.func.distinct(AuditLog.user_id))
    ).filter(AuditLog.created_at >= datetime.utcnow() - timedelta(days=7)).scalar() or 0
    online_cutoff = datetime.utcnow() - timedelta(minutes=30)
    kpi_online = User.query.filter(
        User.last_seen >= online_cutoff, User.is_active == True).count()

    return render_template('admin/audit.html',
                           audit_entries=audit_entries,
                           all_users=all_users,
                           user_f=user_f,
                           action_f=action_f,
                           kpi_total=kpi_total, kpi_today=kpi_today,
                           kpi_this_week=kpi_this_week,
                           kpi_users_active=kpi_users_active,
                           kpi_online=kpi_online)


@admin_bp.route('/audit/export')
@admin_required
def audit_export():
    fmt = request.args.get('fmt', 'csv')
    rows = AuditLog.query.order_by(AuditLog.created_at.desc()).all()
    headers = ['ID', 'Timestamp', 'User', 'Role', 'Action', 'Target', 'IP Address']
    data = [[
        e.id,
        e.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        e.user.full_name if e.user else 'System',
        e.user.role if e.user else '',
        e.action, e.target or '', e.ip_address or ''
    ] for e in rows]
    return _export(headers, data, 'audit_trail', fmt)


@admin_bp.route('/audit/<int:entry_id>/delete', methods=['POST'])
@admin_required
def audit_delete(entry_id):
    entry = db.get_or_404(AuditLog, entry_id)
    db.session.delete(entry)
    _audit('audit.delete', f'Removed entry #{entry_id}')
    db.session.commit()
    flash(f'Audit entry #{entry_id} deleted.', 'success')
    return redirect(request.referrer or url_for('admin.audit_trail'))


@admin_bp.route('/audit/clear', methods=['POST'])
@admin_required
def audit_clear():
    """Delete all audit log entries (irreversible)."""
    count = AuditLog.query.count()
    AuditLog.query.delete()
    _audit('audit.clear', f'Cleared {count} entries')
    db.session.commit()
    flash(f'All {count} audit entries cleared.', 'success')
    return redirect(url_for('admin.audit_trail'))


# ─── SETTINGS ────────────────────────────────────────────────────────────────

@admin_bp.route('/settings')
@admin_required
def settings():
    site = get_site_settings()
    cfg  = current_app.config
    ctx = {
        'mail_from_name':    site.get('MAIL_FROM_NAME',    'MOF Jobs'),
        'mail_from_address': site.get('MAIL_FROM_ADDRESS', cfg.get('MAIL_USERNAME', '')),
        'mail_server':       cfg.get('MAIL_SERVER', ''),
        'mail_port':         cfg.get('MAIL_PORT', 587),
        'mail_username':     cfg.get('MAIL_USERNAME', ''),
    }
    return render_template('admin/settings.html', **ctx)


@admin_bp.route('/settings/save', methods=['POST'])
@admin_required
def settings_save():
    from_name    = request.form.get('mail_from_name', '').strip()
    from_address = request.form.get('mail_from_address', '').strip()
    if not from_name or not from_address:
        flash('From Name and From Address cannot be empty.', 'danger')
        return redirect(url_for('admin.settings'))
    save_site_settings({'MAIL_FROM_NAME': from_name, 'MAIL_FROM_ADDRESS': from_address})
    _audit('settings.email', f'Sender set to {from_name} <{from_address}>')
    db.session.commit()
    flash('Email sender settings saved.', 'success')
    return redirect(url_for('admin.settings'))


@admin_bp.route('/settings/test-email', methods=['POST'])
@admin_required
def test_email():
    """Send a test email to verify SMTP configuration."""
    recipient = request.form.get('recipient', current_user.email).strip()
    try:
        html = render_template('emails/test_email.html', admin=current_user)
        send_email(recipient, '✅ MOF Jobs — Email Configuration Test', html)
        flash(f'Test email sent to {recipient}. Check your inbox.', 'success')
        _audit('email.test', f'To: {recipient}')
        db.session.commit()
    except Exception as ex:
        flash(f'Email failed: {ex}', 'danger')
    return redirect(url_for('admin.settings'))


# ─── COMPANIES ────────────────────────────────────────────────────────────────

INDUSTRY_CHOICES = [
    'Engineering', 'Construction', 'Oil & Gas', 'Manufacturing',
    'Information Technology', 'Finance', 'Healthcare', 'Education',
    'Consulting', 'Government', 'Logistics', 'Real Estate', 'Other',
]
SIZE_CHOICES = ['1–10', '11–50', '51–200', '201–500', '500+']


@admin_bp.route('/companies')
@admin_required
def companies():
    q_str        = request.args.get('q', '').strip()
    industries_f = request.args.getlist('industry')
    status_f     = request.args.get('status', '')
    verified_f   = request.args.get('verified', '')

    q = Company.query
    if q_str:
        q = q.filter(Company.name.ilike(f'%{q_str}%'))
    if industries_f:
        q = q.filter(Company.industry.in_(industries_f))
    if status_f == 'active':
        q = q.filter_by(is_active=True)
    elif status_f == 'inactive':
        q = q.filter_by(is_active=False)
    if verified_f == 'yes':
        q = q.filter_by(is_verified=True)
    elif verified_f == 'no':
        q = q.filter_by(is_verified=False)

    all_companies = q.order_by(Company.name).all()

    kpi_total    = Company.query.count()
    kpi_active   = Company.query.filter_by(is_active=True).count()
    kpi_verified = Company.query.filter_by(is_verified=True).count()
    kpi_open_jobs = Position.query.filter_by(is_active=True).count()
    kpi_followers = db.session.query(db.func.count(CompanyFollow.id)).scalar() or 0
    kpi_new_30d  = Company.query.filter(
        Company.created_at >= datetime.utcnow() - timedelta(days=30)
    ).count() if hasattr(Company, 'created_at') else 0

    return render_template('admin/companies.html', companies=all_companies,
                           q_str=q_str, industries_f=industries_f,
                           status_f=status_f, verified_f=verified_f,
                           INDUSTRY_CHOICES=INDUSTRY_CHOICES,
                           kpi_total=kpi_total, kpi_active=kpi_active,
                           kpi_verified=kpi_verified, kpi_open_jobs=kpi_open_jobs,
                           kpi_followers=kpi_followers, kpi_new_30d=kpi_new_30d)


@admin_bp.route('/companies/export')
@admin_required
def companies_export():
    fmt = request.args.get('fmt', 'csv')
    rows = Company.query.order_by(Company.name).all()
    headers = ['ID', 'Name', 'Industry', 'Size', 'Location', 'Website',
               'Contact Email', 'Status', 'Verified', 'Open Jobs', 'Followers']
    data = [[
        c.id, c.name, c.industry or '', c.size or '', c.location or '',
        c.website or '', c.contact_email or '',
        'Active' if c.is_active else 'Inactive',
        'Yes' if c.is_verified else 'No',
        c.open_jobs_count, c.follower_count
    ] for c in rows]
    return _export(headers, data, 'companies', fmt)


@admin_bp.route('/companies/new', methods=['GET', 'POST'])
@admin_required
def company_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Company name is required.', 'danger')
            return redirect(url_for('admin.company_new'))

        company = Company(
            name          = name,
            description   = request.form.get('description', '').strip(),
            industry      = request.form.get('industry', ''),
            size          = request.form.get('size', ''),
            website       = request.form.get('website', '').strip(),
            location      = request.form.get('location', '').strip(),
            contact_email = request.form.get('contact_email', '').strip(),
            contact_phone = request.form.get('contact_phone', '').strip(),
            is_verified   = bool(request.form.get('is_verified')),
            is_active     = bool(request.form.get('is_active', 'on')),
            created_by    = current_user.id,
        )
        yr = request.form.get('founded_year', '').strip()
        if yr.isdigit():
            company.founded_year = int(yr)
        company.save_slug()

        logo = request.files.get('logo')
        if logo and logo.filename:
            try:
                company.logo_filename = save_company_image(logo)
            except ValueError as e:
                flash(str(e), 'danger')
        cover = request.files.get('cover')
        if cover and cover.filename:
            try:
                company.cover_filename = save_company_image(cover)
            except ValueError as e:
                flash(str(e), 'danger')

        db.session.add(company)
        _audit('company.create', company.name)
        db.session.commit()
        flash(f'Company "{company.name}" created.', 'success')
        return redirect(url_for('admin.company_detail', company_id=company.id))

    return render_template('admin/company_form.html', company=None,
                           industries=INDUSTRY_CHOICES, sizes=SIZE_CHOICES)


@admin_bp.route('/companies/<int:company_id>')
@admin_required
def company_detail(company_id):
    company    = db.get_or_404(Company, company_id)
    managers   = (CompanyMember.query
                  .filter_by(company_id=company_id, role='manager')
                  .all())
    supervisors = User.query.filter_by(role=ROLE_SUPERVISOR, is_active=True).all()
    assigned_ids = {m.user_id for m in managers}
    available_supervisors = [s for s in supervisors if s.id not in assigned_ids]
    positions = company.positions.order_by(Position.created_at.desc()).all()
    return render_template('admin/company_detail.html',
                           company=company,
                           managers=managers,
                           available_supervisors=available_supervisors,
                           positions=positions)


@admin_bp.route('/companies/<int:company_id>/edit', methods=['GET', 'POST'])
@admin_required
def company_edit(company_id):
    company = db.get_or_404(Company, company_id)
    if request.method == 'POST':
        company.name          = request.form.get('name', company.name).strip()
        company.description   = request.form.get('description', '').strip()
        company.industry      = request.form.get('industry', '')
        company.size          = request.form.get('size', '')
        company.website       = request.form.get('website', '').strip()
        company.location      = request.form.get('location', '').strip()
        company.contact_email = request.form.get('contact_email', '').strip()
        company.contact_phone = request.form.get('contact_phone', '').strip()
        company.is_verified   = bool(request.form.get('is_verified'))
        company.is_active     = bool(request.form.get('is_active'))
        yr = request.form.get('founded_year', '').strip()
        company.founded_year  = int(yr) if yr.isdigit() else None

        logo = request.files.get('logo')
        if logo and logo.filename:
            try:
                company.logo_filename = save_company_image(logo)
            except ValueError as e:
                flash(str(e), 'danger')
        cover = request.files.get('cover')
        if cover and cover.filename:
            try:
                company.cover_filename = save_company_image(cover)
            except ValueError as e:
                flash(str(e), 'danger')

        _audit('company.edit', company.name)
        db.session.commit()
        flash('Company updated.', 'success')
        return redirect(url_for('admin.company_detail', company_id=company_id))

    return render_template('admin/company_form.html', company=company,
                           industries=INDUSTRY_CHOICES, sizes=SIZE_CHOICES)


@admin_bp.route('/companies/<int:company_id>/delete', methods=['POST'])
@admin_required
def company_delete(company_id):
    company = db.get_or_404(Company, company_id)
    name = company.name
    db.session.delete(company)
    _audit('company.delete', name)
    db.session.commit()
    flash(f'Company "{name}" deleted.', 'success')
    return redirect(url_for('admin.companies'))


@admin_bp.route('/companies/<int:company_id>/managers/add', methods=['POST'])
@admin_required
def company_manager_add(company_id):
    company = db.get_or_404(Company, company_id)
    user_id = request.form.get('user_id', type=int)
    if not user_id:
        flash('Select a supervisor.', 'danger')
        return redirect(url_for('admin.company_detail', company_id=company_id))
    supervisor = db.get_or_404(User, user_id)
    if supervisor.role != ROLE_SUPERVISOR:
        flash('Selected user is not a supervisor.', 'danger')
        return redirect(url_for('admin.company_detail', company_id=company_id))

    existing = CompanyMember.query.filter_by(
        company_id=company_id, user_id=user_id).first()
    if existing:
        existing.role = 'manager'
    else:
        db.session.add(CompanyMember(
            company_id=company_id, user_id=user_id, role='manager'))

    _audit('company.manager_add', f'{supervisor.full_name} → {company.name}')
    db.session.commit()
    push_notification(
        user_id,
        f'You have been assigned as manager of {company.name}',
        url_for('supervisor.company_dashboard', company_id=company_id),
    )
    db.session.commit()
    try:
        dashboard_url = url_for('supervisor.company_dashboard', company_id=company_id)
        html = render_template(
            'emails/company_manager_assigned.html',
            supervisor=supervisor,
            company=company,
            dashboard_url=dashboard_url,
        )
        send_email(
            supervisor.email,
            f'You have been assigned as manager of {company.name}',
            html,
        )
    except Exception as e:
        current_app.logger.warning(f'Company manager assignment email failed: {e}')
    flash(f'{supervisor.full_name} added as manager of {company.name}.', 'success')
    return redirect(url_for('admin.company_detail', company_id=company_id))


@admin_bp.route('/companies/<int:company_id>/managers/<int:user_id>/remove',
                methods=['POST'])
@admin_required
def company_manager_remove(company_id, user_id):
    company = db.get_or_404(Company, company_id)
    member = CompanyMember.query.filter_by(
        company_id=company_id, user_id=user_id, role='manager').first_or_404()
    db.session.delete(member)
    _audit('company.manager_remove', f'user#{user_id} ← {company.name}')
    db.session.commit()
    flash('Manager removed.', 'success')
    return redirect(url_for('admin.company_detail', company_id=company_id))


# ─── COMPANY JOB ALERTS ───────────────────────────────────────────────────────

def _send_company_job_alerts(position):
    """Email & notify about a new position.
    - Internships: university coordinators + student followers.
    - Regular jobs: all company followers.
    """
    if not position.company_id:
        return
    if position.type == 'Internship':
        _send_internship_alerts(position)
        return
    followers = CompanyFollow.query.filter_by(company_id=position.company_id).all()
    for follow in followers:
        user = follow.user
        try:
            html = render_template(
                'emails/new_job_alert.html',
                user=user, position=position, company=position.company,
            )
            send_email(
                user.email,
                f'New job at {position.company.name}: {position.title}',
                html,
            )
        except Exception as e:
            current_app.logger.warning(
                f'Job alert email failed for {user.email}: {e}')
        push_notification(
            user.id,
            f'New job at {position.company.name}: {position.title}',
            url_for('jobs.detail', job_id=position.id),
            'bi-briefcase-fill',
        )
    if followers:
        db.session.commit()


def _send_internship_alerts(position):
    """Notify all university coordinators + students following the company about a new internship."""
    site_url = current_app.config.get('SITE_URL', '')
    notified_ids = set()

    coord_memberships = UniversityMember.query.filter_by(role='coordinator').all()
    coordinators = User.query.filter(
        User.id.in_([m.user_id for m in coord_memberships]),
        User.is_active == True
    ).all() if coord_memberships else []

    for coord in coordinators:
        push_notification(
            coord.id,
            f'New internship posted: {position.title} at {position.company.name}',
            url_for('jobs.detail', job_id=position.id),
            'bi-mortarboard-fill',
        )
        try:
            html = render_template('emails/new_internship_alert.html',
                recipient=coord, position=position,
                company=position.company, is_coordinator=True, site_url=site_url)
            send_email(
                coord.email,
                f'New internship: {position.title} at {position.company.name}',
                html,
            )
        except Exception as e:
            current_app.logger.warning(f'Internship coordinator alert failed for {coord.email}: {e}')
        notified_ids.add(coord.id)

    student_follows = (
        CompanyFollow.query
        .join(User, User.id == CompanyFollow.user_id)
        .filter(
            CompanyFollow.company_id == position.company_id,
            User.role == ROLE_STUDENT,
            User.is_active == True
        ).all()
    )
    for follow in student_follows:
        student = follow.user
        if student.id in notified_ids:
            continue
        push_notification(
            student.id,
            f'New internship at {position.company.name}: {position.title}',
            url_for('jobs.detail', job_id=position.id),
            'bi-mortarboard-fill',
        )
        try:
            html = render_template('emails/new_internship_alert.html',
                recipient=student, position=position,
                company=position.company, is_coordinator=False, site_url=site_url)
            send_email(
                student.email,
                f'New internship at {position.company.name}: {position.title}',
                html,
            )
        except Exception as e:
            current_app.logger.warning(f'Internship student alert failed for {student.email}: {e}')
        notified_ids.add(student.id)

    if notified_ids:
        db.session.commit()


# ─── SUPERVISOR REQUESTS ──────────────────────────────────────────────────────

@admin_bp.route('/supervisor_requests')
@admin_required
def supervisor_requests():
    status_f = request.args.get('status', '')
    page     = request.args.get('page', 1, type=int)

    q = SupervisorRequest.query
    if status_f:
        q = q.filter_by(status=status_f)

    requests_paged = q.order_by(SupervisorRequest.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False)

    kpi_total    = SupervisorRequest.query.count()
    kpi_pending  = SupervisorRequest.query.filter_by(status='pending').count()
    kpi_approved = SupervisorRequest.query.filter_by(status='approved').count()
    kpi_rejected = SupervisorRequest.query.filter_by(status='rejected').count()

    return render_template('admin/supervisor_requests.html',
                           requests=requests_paged,
                           status_f=status_f,
                           kpi_total=kpi_total,
                           kpi_pending=kpi_pending,
                           kpi_approved=kpi_approved,
                           kpi_rejected=kpi_rejected)


@admin_bp.route('/supervisor_requests/<int:req_id>')
@admin_required
def supervisor_request_detail(req_id):
    req = db.get_or_404(SupervisorRequest, req_id)
    return render_template('admin/supervisor_request_detail.html', req=req)


@admin_bp.route('/supervisor_requests/<int:req_id>/approve', methods=['POST'])
@admin_required
def supervisor_request_approve(req_id):
    req = db.get_or_404(SupervisorRequest, req_id)
    if req.status != 'pending':
        flash('This request is not pending.', 'warning')
        return redirect(url_for('admin.supervisor_request_detail', req_id=req_id))

    # Check if email already taken
    if User.query.filter_by(email=req.email).first():
        flash(f'A user with email {req.email} already exists.', 'danger')
        return redirect(url_for('admin.supervisor_request_detail', req_id=req_id))

    # Create User
    from slugify import slugify as _slugify
    new_user = User(
        full_name     = req.full_name,
        email         = req.email,
        phone         = req.phone,
        role          = ROLE_SUPERVISOR,
        is_active     = True,
        headline      = req.headline,
        bio           = req.bio,
        nationality   = req.nationality,
        location_city = req.location_city,
        gender        = req.gender,
        linkedin_url  = req.linkedin_url,
        password_hash = req.password_hash,  # already hashed
    )
    db.session.add(new_user)
    db.session.flush()  # get new_user.id

    # Create Company
    company = Company(
        name          = req.company_name,
        description   = req.company_description,
        industry      = req.company_industry,
        size          = req.company_size,
        website       = req.company_website,
        location      = req.company_location,
        founded_year  = req.company_founded_year,
        contact_email = req.company_contact_email,
        contact_phone = req.company_contact_phone,
        logo_filename = req.company_logo_filename,
        is_verified   = True,
        is_active     = True,
        created_by    = new_user.id,
    )
    company.save_slug()
    db.session.add(company)
    db.session.flush()

    # Create CompanyMember
    member = CompanyMember(
        company_id = company.id,
        user_id    = new_user.id,
        role       = 'manager',
    )
    db.session.add(member)

    # Mark request approved
    req.status          = 'approved'
    req.reviewed_by_id  = current_user.id
    req.reviewed_at     = datetime.utcnow()

    _audit('supervisor_request.approve',
           f'{req.full_name} <{req.email}> → user#{new_user.id} company#{company.id}')
    db.session.commit()

    # Send welcome email
    try:
        site_url  = current_app.config.get('SITE_URL', '')
        login_url = site_url + url_for('auth.login')
        html = render_template('emails/sup_request_approved.html',
                               req=req, user=new_user, login_url=login_url)
        send_email(new_user.email, 'Your MOF Jobs Supervisor Account is Ready', html)
    except Exception as exc:
        current_app.logger.warning(f'Approval email failed: {exc}')

    flash(f'Approved! User "{new_user.full_name}" and company "{company.name}" created.', 'success')
    return redirect(url_for('admin.supervisor_requests'))


@admin_bp.route('/supervisor_requests/<int:req_id>/reject', methods=['POST'])
@admin_required
def supervisor_request_reject(req_id):
    req = db.get_or_404(SupervisorRequest, req_id)
    if req.status != 'pending':
        flash('This request is not pending.', 'warning')
        return redirect(url_for('admin.supervisor_request_detail', req_id=req_id))

    reason = request.form.get('reason', '').strip()
    req.status           = 'rejected'
    req.rejection_reason = reason or None
    req.reviewed_by_id   = current_user.id
    req.reviewed_at      = datetime.utcnow()

    _audit('supervisor_request.reject', f'{req.full_name} <{req.email}>')
    db.session.commit()

    # Send rejection email with edit link
    try:
        site_url = current_app.config.get('SITE_URL', '')
        edit_url = site_url + url_for('supervisor_apply.apply_edit', token=req.token)
        html = render_template('emails/sup_request_rejected.html',
                               req=req, edit_url=edit_url)
        send_email(req.email, 'Update on Your MOF Jobs Supervisor Application', html)
    except Exception as exc:
        current_app.logger.warning(f'Rejection email failed: {exc}')

    flash(f'Application from "{req.full_name}" rejected.', 'success')
    return redirect(url_for('admin.supervisor_requests'))


@admin_bp.route('/supervisor_requests/<int:req_id>/edit', methods=['POST'])
@admin_required
def supervisor_request_edit(req_id):
    req = db.get_or_404(SupervisorRequest, req_id)

    req.full_name            = request.form.get('full_name', '').strip() or req.full_name
    req.email                = request.form.get('email', '').strip() or req.email
    req.phone                = request.form.get('phone', '').strip() or req.phone
    req.headline             = request.form.get('headline', '').strip() or None
    req.location_city        = request.form.get('location_city', '').strip() or None
    req.nationality          = request.form.get('nationality', '').strip() or None
    req.gender               = request.form.get('gender', '').strip() or None
    req.bio                  = request.form.get('bio', '').strip() or None
    req.linkedin_url         = request.form.get('linkedin_url', '').strip() or None
    req.company_name         = request.form.get('company_name', '').strip() or req.company_name
    req.company_industry     = request.form.get('company_industry', '').strip() or None
    req.company_size         = request.form.get('company_size', '').strip() or None
    req.company_location     = request.form.get('company_location', '').strip() or None
    req.company_website      = request.form.get('company_website', '').strip() or None
    req.company_description  = request.form.get('company_description', '').strip() or None
    req.company_contact_email= request.form.get('company_contact_email', '').strip() or None
    req.company_contact_phone= request.form.get('company_contact_phone', '').strip() or None
    yr = request.form.get('company_founded_year', '').strip()
    req.company_founded_year = int(yr) if yr.isdigit() else req.company_founded_year

    # Optional logo replacement
    logo_file = request.files.get('company_logo')
    if logo_file and logo_file.filename:
        try:
            req.company_logo_filename = save_company_image(logo_file)
        except ValueError as e:
            flash(str(e), 'warning')

    # Optional password reset
    new_pw = request.form.get('new_password', '').strip()
    if new_pw:
        req.set_password(new_pw)

    # Optionally reset to pending so it can be re-reviewed
    if request.form.get('reset_to_pending') and req.status != 'approved':
        req.status = 'pending'
        req.rejection_reason = None

    _audit('supervisor_request.edit', f'{req.full_name} <{req.email}>')
    db.session.commit()
    flash('Application updated successfully.', 'success')
    return redirect(url_for('admin.supervisor_request_detail', req_id=req_id))


# ─── EXPORT HELPER ────────────────────────────────────────────────────────────

def _export(headers, data, filename_base, fmt='csv'):
    """Return a CSV or Excel response for the given data."""
    if fmt == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = filename_base.replace('_', ' ').title()
            # Header row styling
            header_fill = PatternFill('solid', fgColor='0F1923')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            ws.append(headers)
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for row in data:
                ws.append([str(v) if v is not None else '' for v in row])
            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True,
                             download_name=f'{filename_base}.xlsx')
        except ImportError:
            flash('openpyxl is required for Excel export. Install it with: pip install openpyxl', 'warning')
            return redirect(request.referrer or url_for('admin.dashboard'))

    # Default: CSV
    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(headers)
    for row in data:
        writer.writerow([str(v) if v is not None else '' for v in row])
    return Response(
        si.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename_base}.csv'}
    )


# ─────────────────────────────────────────────────────────────────────────────
# UNIVERSITY REQUESTS
# ─────────────────────────────────────────────────────────────────────────────

@admin_bp.route('/university_requests')
@admin_required
def university_requests():
    status_f = request.args.get('status', '')
    page     = request.args.get('page', 1, type=int)
    q = UniversityRequest.query
    if status_f:
        q = q.filter_by(status=status_f)
    reqs = q.order_by(UniversityRequest.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False)
    kpi_total    = UniversityRequest.query.count()
    kpi_pending  = UniversityRequest.query.filter_by(status='pending').count()
    kpi_approved = UniversityRequest.query.filter_by(status='approved').count()
    kpi_rejected = UniversityRequest.query.filter_by(status='rejected').count()
    return render_template('admin/university_requests.html',
        requests=reqs, status_f=status_f,
        kpi_total=kpi_total, kpi_pending=kpi_pending,
        kpi_approved=kpi_approved, kpi_rejected=kpi_rejected)


@admin_bp.route('/university_requests/<int:req_id>')
@admin_required
def university_request_detail(req_id):
    req = UniversityRequest.query.get_or_404(req_id)
    return render_template('admin/university_request_detail.html', req=req)


@admin_bp.route('/university_requests/<int:req_id>/approve', methods=['POST'])
@admin_required
def university_request_approve(req_id):
    req = UniversityRequest.query.get_or_404(req_id)
    if req.status == 'approved':
        flash('Already approved.', 'info')
        return redirect(url_for('admin.university_request_detail', req_id=req_id))

    from slugify import slugify as _slug
    from werkzeug.security import generate_password_hash

    # Create coordinator User
    coordinator = User(
        full_name     = req.full_name,
        email         = req.email,
        phone         = req.phone,
        role          = ROLE_UNIVERSITY_COORD,
        password_hash = req.password_hash,
        headline      = req.headline,
        bio           = req.bio,
        nationality   = req.nationality,
        location_city = req.location_city,
        gender        = req.gender,
        linkedin_url  = req.linkedin_url,
        is_active     = True,
    )
    db.session.add(coordinator)
    db.session.flush()

    # Create University
    base_slug = _slug(req.university_name)
    slug = base_slug
    i = 2
    while University.query.filter_by(slug=slug).first():
        slug = f'{base_slug}-{i}'; i += 1
    univ = University(
        name          = req.university_name,
        slug          = slug,
        description   = req.university_description,
        location      = req.university_location,
        website       = req.university_website,
        contact_email = req.university_contact_email,
        contact_phone = req.university_contact_phone,
        logo_filename = req.university_logo_filename,
        created_by    = current_user.id,
    )
    db.session.add(univ)
    db.session.flush()

    coordinator.university_id = univ.id
    db.session.add(UniversityMember(university_id=univ.id, user_id=coordinator.id, role='coordinator'))

    req.status = 'approved'
    req.reviewed_by_id = current_user.id
    req.reviewed_at = datetime.utcnow()
    db.session.commit()
    _audit('university_request.approve', f'{req.full_name} → {req.university_name}')
    db.session.commit()

    site_url = current_app.config.get('SITE_URL', '')
    try:
        html = render_template('emails/univ_request_approved.html',
                               req=req, login_url=site_url + url_for('auth.login'))
        send_email(req.email,
                   f'Your University Coordinator Account is Approved — {req.university_name}', html)
    except Exception as e:
        current_app.logger.warning(f'Approval email failed: {e}')

    flash(f'Approved — {req.full_name} is now a University Coordinator for {req.university_name}.', 'success')
    return redirect(url_for('admin.university_request_detail', req_id=req_id))


@admin_bp.route('/university_requests/<int:req_id>/reject', methods=['POST'])
@admin_required
def university_request_reject(req_id):
    req = UniversityRequest.query.get_or_404(req_id)
    reason = request.form.get('rejection_reason', '').strip()
    req.status = 'rejected'
    req.rejection_reason = reason or None
    req.reviewed_by_id = current_user.id
    req.reviewed_at = datetime.utcnow()
    db.session.commit()
    _audit('university_request.reject', f'{req.full_name}')
    db.session.commit()

    site_url = current_app.config.get('SITE_URL', '')
    edit_url = site_url + url_for('university_apply.apply_edit', token=req.token)
    try:
        html = render_template('emails/univ_request_rejected.html',
                               req=req, edit_url=edit_url)
        send_email(req.email,
                   f'University Coordinator Application — Update Required', html)
    except Exception as e:
        current_app.logger.warning(f'Rejection email failed: {e}')

    flash('Application rejected. Applicant has been notified.', 'success')
    return redirect(url_for('admin.university_request_detail', req_id=req_id))


@admin_bp.route('/university_requests/<int:req_id>/edit', methods=['POST'])
@admin_required
def university_request_edit(req_id):
    req = UniversityRequest.query.get_or_404(req_id)
    req.full_name                = request.form.get('full_name', '').strip() or req.full_name
    req.email                    = request.form.get('email', '').strip() or req.email
    req.phone                    = request.form.get('phone', '').strip() or req.phone
    req.headline                 = request.form.get('headline', '').strip() or req.headline
    req.nationality              = request.form.get('nationality', '').strip() or req.nationality
    req.location_city            = request.form.get('location_city', '').strip() or req.location_city
    req.gender                   = request.form.get('gender', '').strip() or req.gender
    req.linkedin_url             = request.form.get('linkedin_url', '').strip() or req.linkedin_url
    bio = request.form.get('bio', '').strip()
    if bio:
        req.bio = bio
    req.university_name          = request.form.get('university_name', '').strip() or req.university_name
    req.university_location      = request.form.get('university_location', '').strip() or req.university_location
    req.university_website       = request.form.get('university_website', '').strip() or req.university_website
    req.university_contact_email = request.form.get('university_contact_email', '').strip() or req.university_contact_email
    req.university_contact_phone = request.form.get('university_contact_phone', '').strip() or req.university_contact_phone
    univ_desc = request.form.get('university_description', '').strip()
    if univ_desc:
        req.university_description = univ_desc
    logo_file = request.files.get('university_logo')
    if logo_file and logo_file.filename:
        try:
            req.university_logo_filename = save_company_image(logo_file)
        except ValueError as e:
            flash(str(e), 'warning')
    new_pw = request.form.get('new_password', '').strip()
    if new_pw:
        req.set_password(new_pw)
    if request.form.get('reset_to_pending') and req.status != 'approved':
        req.status = 'pending'
    _audit('university_request.edit', f'#{req_id} {req.full_name}')
    db.session.commit()
    flash('Request updated.', 'success')
    return redirect(url_for('admin.university_request_detail', req_id=req_id))


# ─────────────────────────────────────────────────────────────────────────────
# UNIVERSITIES
# ─────────────────────────────────────────────────────────────────────────────

@admin_bp.route('/coordinators')
@admin_required
def coordinators():
    """List all university coordinators with their universities."""
    q_str = request.args.get('q', '').strip()
    univ_f = request.args.get('university', type=int)
    view = request.args.get('view', 'cards')
    if view not in ('cards', 'list'):
        view = 'cards'

    # Base query: every coordinator user (active or not).
    base = User.query.filter(User.role == ROLE_UNIVERSITY_COORD)
    if q_str:
        like = f'%{q_str}%'
        base = base.filter(or_(User.full_name.ilike(like),
                               User.email.ilike(like),
                               User.phone.ilike(like)))
    coords_users = base.order_by(User.full_name).all()

    # Map user_id -> list of (UniversityMember, University) so we can show all
    # university links for each coordinator (a coord may belong to >1 univ).
    memberships = (UniversityMember.query
                   .filter_by(role='coordinator')
                   .join(User, UniversityMember.user_id == User.id)
                   .filter(User.role == ROLE_UNIVERSITY_COORD)
                   .all())
    by_user = {}
    for m in memberships:
        by_user.setdefault(m.user_id, []).append(m)

    rows = []
    for u in coords_users:
        ms = by_user.get(u.id, [])
        if univ_f:
            ms = [m for m in ms if m.university_id == univ_f]
            if not ms:
                continue
        rows.append({'user': u, 'memberships': ms})

    # Sidebar filter options
    universities_list = University.query.order_by(University.name).all()

    return render_template('admin/coordinators.html',
                           rows=rows, view=view, q=q_str,
                           univ_f=univ_f, universities_list=universities_list,
                           total=len(rows))


def _csv_response(filename, header, rows):
    """Build a UTF-8 CSV Response with BOM (so Excel opens it correctly)."""
    buf = io.StringIO()
    buf.write('\ufeff')  # BOM for Excel
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow(header)
    for r in rows:
        w.writerow(['' if v is None else v for v in r])
    resp = Response(buf.getvalue(), mimetype='text/csv; charset=utf-8')
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@admin_bp.route('/coordinators/export.csv')
@admin_required
def coordinators_export():
    """Export the coordinators directory (respects q + university filters)."""
    q_str = request.args.get('q', '').strip()
    univ_f = request.args.get('university', type=int)

    base = User.query.filter(User.role == ROLE_UNIVERSITY_COORD)
    if q_str:
        like = f'%{q_str}%'
        base = base.filter(or_(User.full_name.ilike(like),
                               User.email.ilike(like),
                               User.phone.ilike(like)))
    coords_users = base.order_by(User.full_name).all()

    memberships = (UniversityMember.query
                   .filter_by(role='coordinator')
                   .join(User, UniversityMember.user_id == User.id)
                   .filter(User.role == ROLE_UNIVERSITY_COORD)
                   .all())
    by_user = {}
    for m in memberships:
        by_user.setdefault(m.user_id, []).append(m)

    header = ['Coordinator ID', 'Full Name', 'Email', 'Phone', 'Active',
              'University', 'Department', 'Class Scope', 'Joined At',
              'Last Login', 'Bio']
    rows = []
    for u in coords_users:
        ms = by_user.get(u.id, [])
        if univ_f:
            ms = [m for m in ms if m.university_id == univ_f]
            if not ms:
                continue
        if not ms:
            rows.append([u.id, u.full_name, u.email, u.phone or '',
                         'Yes' if u.is_active else 'No',
                         '', '', '', '',
                         u.last_login.isoformat(sep=' ', timespec='minutes') if u.last_login else '',
                         (u.bio or '').replace('\r', ' ').replace('\n', ' ')])
            continue
        for m in ms:
            uni = m.university.name if m.university else ''
            dept = m.department.name if getattr(m, 'department', None) else ''
            cls = getattr(m, 'class_scope', '') or ''
            joined = m.joined_at.isoformat(sep=' ', timespec='minutes') if getattr(m, 'joined_at', None) else ''
            rows.append([u.id, u.full_name, u.email, u.phone or '',
                         'Yes' if u.is_active else 'No',
                         uni, dept, cls, joined,
                         u.last_login.isoformat(sep=' ', timespec='minutes') if u.last_login else '',
                         (u.bio or '').replace('\r', ' ').replace('\n', ' ')])

    ts = datetime.utcnow().strftime('%Y%m%d-%H%M')
    return _csv_response(f'coordinators-{ts}.csv', header, rows)


@admin_bp.route('/coordinators/<int:user_id>')
@admin_required
def coordinator_detail(user_id):
    """Detail page for one coordinator: profile + supervised students."""
    coord = db.get_or_404(User, user_id)
    if coord.role != ROLE_UNIVERSITY_COORD:
        flash('That user is not a university coordinator.', 'warning')
        return redirect(url_for('admin.coordinators'))

    view = request.args.get('view', 'cards')
    if view not in ('cards', 'list'):
        view = 'cards'
    q_str = request.args.get('q', '').strip()

    # Memberships of this coordinator (a coord may belong to >1 university,
    # and may be scoped to a department within each).
    memberships = (UniversityMember.query
                   .filter_by(user_id=coord.id, role='coordinator')
                   .all())

    # Build the supervised-students set:
    #  - take every (university_id, department_id) pair the coord covers
    #  - if department_id is None, the coord supervises ALL students of that university
    #  - otherwise only students whose university_department_id matches
    student_q = User.query.filter(
        User.role == ROLE_STUDENT,
        User.is_active == True,
    )

    if not memberships:
        students = []
    else:
        # OR together per-membership scope filters.
        scope_filters = []
        for m in memberships:
            if m.department_id:
                scope_filters.append(and_(
                    User.university_id == m.university_id,
                    User.university_department_id == m.department_id,
                ))
            else:
                scope_filters.append(User.university_id == m.university_id)

        student_q = student_q.filter(or_(*scope_filters))

        if q_str:
            like = f'%{q_str}%'
            student_q = student_q.filter(or_(
                User.full_name.ilike(like),
                User.email.ilike(like),
                User.university_major.ilike(like),
                User.university_class.ilike(like),
                User.student_id_number.ilike(like),
            ))

        students = student_q.order_by(User.full_name).all()

    # Application stats per student (under positions of type=Internship)
    student_ids = [s.id for s in students]
    apps_by_student = {}
    if student_ids:
        all_apps = (Application.query
                    .filter(Application.applicant_id.in_(student_ids))
                    .all())
        for a in all_apps:
            apps_by_student.setdefault(a.applicant_id, []).append(a)

    # Pre-build a lightweight stats dict
    stats_by_student = {}
    for sid in student_ids:
        apps = apps_by_student.get(sid, [])
        stats_by_student[sid] = {
            'total': len(apps),
            'hired': sum(1 for a in apps if a.status == 'Hired'),
            'pending': sum(1 for a in apps if a.status not in ('Hired', 'Rejected')),
        }

    return render_template('admin/coordinator_detail.html',
                           coord=coord,
                           memberships=memberships,
                           students=students,
                           stats=stats_by_student,
                           view=view, q=q_str)


@admin_bp.route('/coordinators/<int:user_id>/export.csv')
@admin_required
def coordinator_detail_export(user_id):
    """Export the supervised-students list for one coordinator (with stats)."""
    coord = db.get_or_404(User, user_id)
    if coord.role != ROLE_UNIVERSITY_COORD:
        abort(404)

    q_str = request.args.get('q', '').strip()
    memberships = (UniversityMember.query
                   .filter_by(user_id=coord.id, role='coordinator')
                   .all())

    students = []
    if memberships:
        student_q = User.query.filter(User.role == ROLE_STUDENT, User.is_active == True)
        scope_filters = []
        for m in memberships:
            if m.department_id:
                scope_filters.append(and_(
                    User.university_id == m.university_id,
                    User.university_department_id == m.department_id,
                ))
            else:
                scope_filters.append(User.university_id == m.university_id)
        student_q = student_q.filter(or_(*scope_filters))
        if q_str:
            like = f'%{q_str}%'
            student_q = student_q.filter(or_(
                User.full_name.ilike(like),
                User.email.ilike(like),
                User.university_major.ilike(like),
                User.university_class.ilike(like),
                User.student_id_number.ilike(like),
            ))
        students = student_q.order_by(User.full_name).all()

    student_ids = [s.id for s in students]
    apps_by_student = {}
    if student_ids:
        for a in Application.query.filter(Application.applicant_id.in_(student_ids)).all():
            apps_by_student.setdefault(a.applicant_id, []).append(a)

    header = ['Student ID', 'Full Name', 'Email', 'Phone', 'Student Number',
              'University', 'Department', 'Major', 'Class',
              'Graduation Year', 'Active', 'Created At', 'Last Login',
              'Total Applications', 'Pending', 'Hired', 'Rejected']
    rows = []
    for s in students:
        apps = apps_by_student.get(s.id, [])
        total = len(apps)
        hired = sum(1 for a in apps if a.status == 'Hired')
        rejected = sum(1 for a in apps if a.status == 'Rejected')
        pending = sum(1 for a in apps if a.status not in ('Hired', 'Rejected'))
        uni = s.university.name if getattr(s, 'university', None) else (s.university_name or '')
        dept = s.university_department.name if getattr(s, 'university_department', None) else ''
        rows.append([
            s.id, s.full_name, s.email, s.phone or '', s.student_id_number or '',
            uni, dept, s.university_major or '', s.university_class or '',
            s.graduation_year or '',
            'Yes' if s.is_active else 'No',
            s.created_at.isoformat(sep=' ', timespec='minutes') if s.created_at else '',
            s.last_login.isoformat(sep=' ', timespec='minutes') if s.last_login else '',
            total, pending, hired, rejected,
        ])

    safe_name = ''.join(c for c in (coord.full_name or f'coord-{coord.id}') if c.isalnum() or c in ('-', '_', ' ')).strip().replace(' ', '_') or f'coord-{coord.id}'
    ts = datetime.utcnow().strftime('%Y%m%d-%H%M')
    return _csv_response(f'{safe_name}-students-{ts}.csv', header, rows)


@admin_bp.route('/universities')
@admin_required
def universities():
    q_str    = request.args.get('q', '').strip()
    status_f = request.args.get('status', '')
    page     = request.args.get('page', 1, type=int)

    q = University.query
    if q_str:
        q = q.filter(University.name.ilike(f'%{q_str}%'))
    if status_f == 'active':
        q = q.filter_by(is_active=True)
    elif status_f == 'inactive':
        q = q.filter_by(is_active=False)

    univs = q.order_by(University.name).paginate(page=page, per_page=20, error_out=False)

    kpi_total    = University.query.count()
    kpi_active   = University.query.filter_by(is_active=True).count()
    kpi_students = User.query.filter_by(role=ROLE_STUDENT).count()
    kpi_coords   = User.query.filter_by(role=ROLE_UNIVERSITY_COORD).count()

    return render_template('admin/universities.html',
        univs=univs, q_str=q_str, status_f=status_f,
        kpi_total=kpi_total, kpi_active=kpi_active,
        kpi_students=kpi_students, kpi_coords=kpi_coords)


@admin_bp.route('/universities/<int:univ_id>')
@admin_required
def university_detail(univ_id):
    univ = db.get_or_404(University, univ_id)
    page = request.args.get('page', 1, type=int)
    q_str = request.args.get('q', '').strip()
    department_id = request.args.get('department_id', type=int)
    class_filter = request.args.get('class_filter', '').strip()
    major_filter = request.args.get('major_filter', '').strip()
    graduation_year = request.args.get('graduation_year', type=int)
    departments = (UniversityDepartment.query
                   .filter_by(university_id=univ_id)
                   .order_by(UniversityDepartment.college.asc(), UniversityDepartment.name.asc())
                   .all())
    coords = (UniversityMember.query
              .filter_by(university_id=univ_id)
              .all())
    students_q = _admin_university_students_query(
        univ_id,
        q_str,
        department_id,
        class_filter,
        major_filter,
        graduation_year,
    )
    students = students_q.order_by(User.full_name.asc()).paginate(page=page, per_page=20, error_out=False)
    student_total = User.query.filter_by(university_id=univ_id, role=ROLE_STUDENT).count()
    available_coords = User.query.filter_by(role=ROLE_UNIVERSITY_COORD, is_active=True).all()
    assigned_ids = {m.user_id for m in coords}
    available_coords = [c for c in available_coords if c.id not in assigned_ids]
    assignable_students = (User.query
                           .filter_by(role=ROLE_STUDENT)
                           .filter(or_(User.university_id.is_(None), User.university_id != univ_id))
                           .order_by(User.full_name.asc())
                           .limit(100)
                           .all())

    internship_count = (Application.query
        .join(Application.position)
        .join(User, Application.applicant_id == User.id)
        .filter(User.university_id == univ_id, Position.type == 'Internship')
        .count())

    return render_template('admin/university_detail.html',
        univ=univ,
        departments=departments,
        coords=coords,
        students=students,
        student_total=student_total,
        q_str=q_str,
        department_filter=department_id,
        class_filter=class_filter,
        major_filter=major_filter,
        graduation_year_filter=graduation_year,
        assignable_students=assignable_students,
        available_coords=available_coords,
        internship_count=internship_count,
        edit_open=(request.args.get('edit') in ('1', 'true', 'yes')))


@admin_bp.route('/universities/<int:univ_id>/edit', methods=['GET', 'POST'])
@admin_required
def university_edit(univ_id):
    univ = db.get_or_404(University, univ_id)
    if request.method == 'GET':
        # Unified entry point: send admins to the detail page with the inline
        # edit panel auto-opened.
        return redirect(url_for('admin.university_detail', univ_id=univ_id, edit=1) + '#tab-info')
    univ.name          = request.form.get('name', univ.name).strip()
    univ.description   = request.form.get('description', '').strip() or None
    univ.location      = request.form.get('location', '').strip() or None
    univ.website       = request.form.get('website', '').strip() or None
    univ.contact_email = request.form.get('contact_email', '').strip() or None
    univ.contact_phone = request.form.get('contact_phone', '').strip() or None
    univ.is_active     = bool(request.form.get('is_active'))

    logo = request.files.get('logo')
    if logo and logo.filename:
        try:
            univ.logo_filename = save_company_image(logo)
        except ValueError as e:
            flash(str(e), 'warning')

    banner = request.files.get('banner')
    if banner and banner.filename:
        try:
            univ.banner_filename = save_company_image(banner)
        except ValueError as e:
            flash(str(e), 'warning')

    _audit('university.edit', univ.name)
    db.session.commit()
    flash('University updated.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/verify', methods=['POST'])
@admin_required
def university_verify(univ_id):
    """Admin marks a university as verified — locks coordinator out of edits."""
    univ = db.get_or_404(University, univ_id)
    action = (request.form.get('action') or 'verify').lower()
    if action == 'unverify':
        univ.is_verified = False
        univ.verified_at = None
        univ.verified_by_id = None
        _audit('university.unverify', univ.name)
        flash(f'{univ.name} marked as unverified — coordinator can edit again.', 'info')
    else:
        univ.is_verified = True
        univ.verified_at = datetime.utcnow()
        univ.verified_by_id = current_user.id
        _audit('university.verify', univ.name)
        flash(f'{univ.name} verified — coordinator edit access is now locked.', 'success')
    db.session.commit()
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/delete', methods=['POST'])
@admin_required
def university_delete(univ_id):
    univ = db.get_or_404(University, univ_id)
    name = univ.name
    # Unlink students
    User.query.filter_by(university_id=univ_id).update({
        'university_id': None,
        'university_department_id': None,
        'university_class': None,
        'university_name': None,
    })
    db.session.delete(univ)
    _audit('university.delete', name)
    db.session.commit()
    flash(f'University "{name}" deleted.', 'success')
    return redirect(url_for('admin.universities'))


@admin_bp.route('/universities/new', methods=['GET', 'POST'])
@admin_required
def university_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('University name is required.', 'danger')
            return redirect(url_for('admin.university_new'))
        univ = University(
            name          = name,
            description   = request.form.get('description', '').strip() or None,
            location      = request.form.get('location', '').strip() or None,
            website       = request.form.get('website', '').strip() or None,
            contact_email = request.form.get('contact_email', '').strip() or None,
            contact_phone = request.form.get('contact_phone', '').strip() or None,
            is_active     = bool(request.form.get('is_active', True)),
            created_by    = current_user.id,
        )
        univ.save_slug()
        logo = request.files.get('logo')
        if logo and logo.filename:
            try:
                univ.logo_filename = save_company_image(logo)
            except ValueError as e:
                flash(str(e), 'warning')
        db.session.add(univ)
        _audit('university.create', univ.name)
        db.session.commit()
        flash(f'University "{univ.name}" created.', 'success')
        return redirect(url_for('admin.university_detail', univ_id=univ.id))
    return render_template('admin/university_form.html', univ=None)


@admin_bp.route('/universities/<int:univ_id>/departments/add', methods=['POST'])
@admin_required
def university_department_add(univ_id):
    univ = db.get_or_404(University, univ_id)
    name = request.form.get('name', '').strip()
    college = request.form.get('college', '').strip() or None
    if not name:
        flash('Department name is required.', 'danger')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    exists = UniversityDepartment.query.filter_by(university_id=univ_id, name=name).first()
    if exists:
        flash('This department already exists for the university.', 'warning')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    db.session.add(UniversityDepartment(
        university_id=univ_id,
        name=name,
        college=college,
        is_active=True,
    ))
    _audit('university.department_add', f'{univ.name}: {name}')
    db.session.commit()
    flash('Department added.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/departments/<int:dept_id>/edit', methods=['POST'])
@admin_required
def university_department_edit(univ_id, dept_id):
    univ = db.get_or_404(University, univ_id)
    dept = UniversityDepartment.query.filter_by(id=dept_id, university_id=univ_id).first_or_404()
    name = request.form.get('name', '').strip()
    college = request.form.get('college', '').strip() or None

    if not name:
        flash('Department name is required.', 'danger')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    exists = (UniversityDepartment.query
              .filter(UniversityDepartment.university_id == univ_id,
                      UniversityDepartment.name == name,
                      UniversityDepartment.id != dept_id)
              .first())
    if exists:
        flash('Another department with this name already exists.', 'warning')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    dept.name = name
    dept.college = college
    _audit('university.department_edit', f'{univ.name}: {dept_id}')
    db.session.commit()
    flash('Department updated.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/departments/<int:dept_id>/delete', methods=['POST'])
@admin_required
def university_department_delete(univ_id, dept_id):
    univ = db.get_or_404(University, univ_id)
    dept = UniversityDepartment.query.filter_by(id=dept_id, university_id=univ_id).first_or_404()

    assigned_coords = UniversityMember.query.filter_by(university_id=univ_id, department_id=dept_id).count()
    assigned_students = User.query.filter_by(university_id=univ_id, university_department_id=dept_id, role=ROLE_STUDENT).count()
    if assigned_coords or assigned_students:
        flash('Cannot delete department while coordinators or students are assigned to it.', 'warning')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    db.session.delete(dept)
    _audit('university.department_delete', f'{univ.name}: {dept.name}')
    db.session.commit()
    flash('Department deleted.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/coordinators/add', methods=['POST'])
@admin_required
def university_coordinator_add(univ_id):
    univ = db.get_or_404(University, univ_id)
    user_id = request.form.get('user_id', type=int)
    department_id = request.form.get('department_id', type=int)
    class_scope = request.form.get('class_scope', '').strip() or None

    if not user_id:
        flash('Select a coordinator.', 'danger')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    if department_id:
        dept = UniversityDepartment.query.filter_by(id=department_id, university_id=univ_id).first()
        if not dept:
            flash('Invalid department selected.', 'danger')
            return redirect(url_for('admin.university_detail', univ_id=univ_id))

    coord = db.get_or_404(User, user_id)
    existing = UniversityMember.query.filter_by(university_id=univ_id, user_id=user_id).first()
    if existing:
        existing.department_id = department_id
        existing.class_scope = class_scope
    else:
        db.session.add(UniversityMember(
            university_id=univ_id,
            user_id=user_id,
            role='coordinator',
            department_id=department_id,
            class_scope=class_scope,
        ))

    coord.university_id = univ_id
    _audit('university.coordinator_add', f'{coord.full_name} → {univ.name}')
    db.session.commit()
    flash(f'{coord.full_name} added as coordinator.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/coordinators/<int:user_id>/scope', methods=['POST'])
@admin_required
def university_coordinator_scope_update(univ_id, user_id):
    univ = db.get_or_404(University, univ_id)
    member = UniversityMember.query.filter_by(university_id=univ_id, user_id=user_id).first_or_404()
    department_id = request.form.get('department_id', type=int)
    class_scope = request.form.get('class_scope', '').strip() or None

    if department_id:
        dept = UniversityDepartment.query.filter_by(id=department_id, university_id=univ_id).first()
        if not dept:
            flash('Invalid department selected.', 'danger')
            return redirect(url_for('admin.university_detail', univ_id=univ_id))

    member.department_id = department_id
    member.class_scope = class_scope
    _audit('university.coordinator_scope_update', f'user#{user_id} @ {univ.name}')
    db.session.commit()
    flash('Coordinator scope updated.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/coordinators/<int:user_id>/remove',
                methods=['POST'])
@admin_required
def university_coordinator_remove(univ_id, user_id):
    univ = db.get_or_404(University, univ_id)
    member = UniversityMember.query.filter_by(
        university_id=univ_id, user_id=user_id).first_or_404()
    db.session.delete(member)

    remaining = UniversityMember.query.filter_by(user_id=user_id).count()
    if remaining == 0:
        user = User.query.get(user_id)
        if user:
            user.university_id = None
            user.university_department_id = None
            user.university_class = None

    _audit('university.coordinator_remove', f'user#{user_id} ← {univ.name}')
    db.session.commit()
    flash('Coordinator removed.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/students/export')
@admin_required
def university_students_export(univ_id):
    db.get_or_404(University, univ_id)
    fmt = request.args.get('fmt', 'xlsx').lower()
    q_str = request.args.get('q', '').strip()
    department_id = request.args.get('department_id', type=int)
    class_filter = request.args.get('class_filter', '').strip()
    major_filter = request.args.get('major_filter', '').strip()
    graduation_year = request.args.get('graduation_year', type=int)
    rows = (_admin_university_students_query(univ_id, q_str, department_id, class_filter, major_filter, graduation_year)
            .order_by(User.full_name.asc())
            .all())

    headers = [label for _, label in _ADMIN_UNIVERSITY_STUDENT_COLS]
    data = []
    for student in rows:
        data.append([
            student.student_id_number,
            student.full_name,
            student.email,
            student.phone,
            student.university_department.full_name() if student.university_department else '',
            student.university_class,
            student.university_major,
            student.graduation_year,
        ])
    return _export(headers, data, f'university_{univ_id}_students', fmt)


@admin_bp.route('/universities/<int:univ_id>/students/import-template')
@admin_required
def university_students_import_template(univ_id):
    db.get_or_404(University, univ_id)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    header_fill = PatternFill('solid', fgColor='1a5c38')
    header_font = Font(color='FFFFFF', bold=True)
    example_data = {
        'student_id_number': '20231001',
        'full_name': 'Ahmed Al-Rashidi',
        'email': 'ahmed@example.com',
        'phone': '+9647501234567',
        'department_name': 'Engineering',
        'university_class': '2027',
        'university_major': 'Computer Engineering',
        'graduation_year': '2027',
    }
    for ci, (attr, label) in enumerate(_ADMIN_UNIVERSITY_STUDENT_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(label) + 4, 18)
        ws.cell(row=2, column=ci, value=example_data.get(attr, ''))

    example_fill = PatternFill('solid', fgColor='F3F4F6')
    for ci in range(1, len(_ADMIN_UNIVERSITY_STUDENT_COLS) + 1):
        ws.cell(row=2, column=ci).fill = example_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='students_import_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/universities/<int:univ_id>/students/import', methods=['POST'])
@admin_required
def university_students_import(univ_id):
    univ = db.get_or_404(University, univ_id)
    import openpyxl

    upload = request.files.get('file')
    default_department_id = request.form.get('default_department_id', type=int)
    default_class = request.form.get('default_class', '').strip() or None

    if default_department_id and not _resolve_university_department(univ_id, department_id=default_department_id):
        flash('Invalid default department selected.', 'danger')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))
    if not upload or not upload.filename:
        flash('No file selected.', 'danger')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))
    if not upload.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(upload.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash('Could not read the uploaded file. Make sure it is a valid Excel workbook.', 'danger')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    headers = [str(cell.value or '').strip() for cell in ws[1]]
    label_to_attr = {label: attr for attr, label in _ADMIN_UNIVERSITY_STUDENT_COLS}
    col_map = {}
    for ci, header in enumerate(headers):
        if header in label_to_attr:
            col_map[ci] = label_to_attr[header]

    if 'email' not in col_map.values() or 'full_name' not in col_map.values():
        flash('Import failed: the file must have "Email" and "Full Name" columns.', 'danger')
        return redirect(url_for('admin.university_detail', univ_id=univ_id))

    created = linked = skipped = 0
    departments = UniversityDepartment.query.filter_by(university_id=univ_id).all()
    department_lookup = {}
    for department in departments:
        department_lookup[_normalize_text(department.name)] = department
        department_lookup[_normalize_text(department.full_name())] = department

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue

        data = {}
        for ci, attr in col_map.items():
            value = row[ci] if ci < len(row) else None
            data[attr] = str(value).strip() if value is not None else ''

        email = data.get('email', '').lower()
        full_name = data.get('full_name', '').strip()
        if not email or not full_name:
            skipped += 1
            continue

        department = None
        if data.get('department_name'):
            department = department_lookup.get(_normalize_text(data['department_name']))
            if not department:
                skipped += 1
                continue
        elif default_department_id:
            department = _resolve_university_department(univ_id, department_id=default_department_id)

        class_value = data.get('university_class', '').strip() or default_class
        existing = User.query.filter_by(email=email).first()
        if existing:
            if existing.role != ROLE_STUDENT:
                skipped += 1
                continue
            existing.university_id = univ.id
            existing.university_name = univ.name
            existing.university_department_id = department.id if department else None
            existing.university_class = class_value
            for attr in ('phone', 'university_major', 'student_id_number'):
                if data.get(attr):
                    setattr(existing, attr, data[attr])
            if data.get('graduation_year'):
                existing.graduation_year = _parse_int(data['graduation_year'])
            linked += 1
            continue

        temp_pw = secrets.token_urlsafe(10)
        student = User(
            full_name=full_name,
            email=email,
            phone=data.get('phone') or None,
            role=ROLE_STUDENT,
            is_active=True,
            university_id=univ.id,
            university_name=univ.name,
            university_department_id=department.id if department else None,
            university_class=class_value,
            university_major=data.get('university_major') or None,
            graduation_year=_parse_int(data.get('graduation_year', '')),
            student_id_number=data.get('student_id_number') or None,
        )
        student.set_password(temp_pw)
        db.session.add(student)
        created += 1

    _audit('university.students_import', f'{univ.name} created={created} linked={linked} skipped={skipped}')
    db.session.commit()
    flash(f'Import complete. Created: {created}, linked: {linked}, skipped: {skipped}.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/students/assign', methods=['POST'])
@admin_required
def university_student_assign(univ_id):
    univ = db.get_or_404(University, univ_id)
    student_id = request.form.get('student_id', type=int)
    email = request.form.get('email', '').strip().lower()
    full_name = request.form.get('full_name', '').strip()
    department_id = request.form.get('department_id', type=int)
    class_value = request.form.get('university_class', '').strip() or None
    major = request.form.get('university_major', '').strip() or None
    graduation_year = _parse_int(request.form.get('graduation_year', ''))

    department = None
    if department_id:
        department = _resolve_university_department(univ_id, department_id=department_id)
        if not department:
            flash('Invalid department selected.', 'danger')
            return redirect(url_for('admin.university_detail', univ_id=univ_id))

    student = None
    if student_id:
        student = User.query.filter_by(id=student_id, role=ROLE_STUDENT).first()
    elif email:
        student = User.query.filter_by(email=email).first()
        if student and student.role != ROLE_STUDENT:
            flash('That email belongs to a non-student account.', 'danger')
            return redirect(url_for('admin.university_detail', univ_id=univ_id))

    if student is None:
        if not email or not full_name:
            flash('Select a student or provide both full name and email to create one.', 'danger')
            return redirect(url_for('admin.university_detail', univ_id=univ_id))
        temp_pw = secrets.token_urlsafe(10)
        student = User(
            full_name=full_name,
            email=email,
            role=ROLE_STUDENT,
            is_active=True,
            university_major=major,
        )
        student.set_password(temp_pw)
        db.session.add(student)

    student.university_id = univ.id
    student.university_name = univ.name
    student.university_department_id = department.id if department else None
    student.university_class = class_value
    if major:
        student.university_major = major
    if graduation_year:
        student.graduation_year = graduation_year

    _audit('university.student_assign', f'{student.email} → {univ.name}')
    db.session.commit()
    flash(f'{student.full_name} assigned to {univ.name}.', 'success')
    return redirect(url_for('admin.university_detail', univ_id=univ_id))


@admin_bp.route('/universities/<int:univ_id>/students/<int:student_id>/update', methods=['POST'])
@admin_required
def university_student_update(univ_id, student_id):
    univ = db.get_or_404(University, univ_id)
    student = User.query.filter_by(id=student_id, university_id=univ_id, role=ROLE_STUDENT).first_or_404()
    department_id = request.form.get('department_id', type=int)
    class_value = request.form.get('university_class', '').strip() or None
    major = request.form.get('university_major', '').strip() or None
    graduation_year_raw = request.form.get('graduation_year', '').strip()

    department = None
    if department_id:
        department = _resolve_university_department(univ_id, department_id=department_id)
        if not department:
            flash('Invalid department selected.', 'danger')
            return redirect(url_for('admin.university_detail', univ_id=univ_id))

    student.university_department_id = department.id if department else None
    student.university_class = class_value
    student.university_major = major
    student.graduation_year = _parse_int(graduation_year_raw)
    _audit('university.student_update', f'{student.email} @ {univ.name}')
    db.session.commit()
    flash('Student assignment updated.', 'success')
    return _admin_university_student_redirect(univ_id)


@admin_bp.route('/universities/<int:univ_id>/students/bulk', methods=['POST'])
@admin_required
def university_students_bulk(univ_id):
    univ = db.get_or_404(University, univ_id)
    action = request.form.get('bulk_action', '').strip()
    department_id = request.form.get('bulk_department_id', type=int)
    class_value = request.form.get('bulk_class', '').strip() or None
    student_ids = request.form.getlist('student_ids')

    if not student_ids:
        flash('Select at least one student.', 'warning')
        return _admin_university_student_redirect(univ_id)

    students = (User.query
                .filter(User.id.in_(student_ids), User.university_id == univ_id, User.role == ROLE_STUDENT)
                .all())
    if not students:
        flash('No matching students found for bulk action.', 'warning')
        return _admin_university_student_redirect(univ_id)

    department = None
    if department_id:
        department = _resolve_university_department(univ_id, department_id=department_id)
        if not department:
            flash('Invalid bulk department selected.', 'danger')
            return _admin_university_student_redirect(univ_id)

    if action == 'set_department':
        for student in students:
            student.university_department_id = department.id if department else None
            if class_value is not None:
                student.university_class = class_value
        message = f'Updated {len(students)} student(s).'
    elif action == 'clear_department':
        for student in students:
            student.university_department_id = None
            if class_value is not None:
                student.university_class = class_value
        message = f'Cleared department for {len(students)} student(s).'
    elif action == 'unlink':
        for student in students:
            student.university_id = None
            student.university_name = None
            student.university_department_id = None
            student.university_class = None
        message = f'Unlinked {len(students)} student(s).'
    else:
        flash('Choose a valid bulk action.', 'warning')
        return _admin_university_student_redirect(univ_id)

    _audit('university.students_bulk', f'{univ.name} action={action} count={len(students)}')
    db.session.commit()
    flash(message, 'success')
    return _admin_university_student_redirect(univ_id)


@admin_bp.route('/universities/<int:univ_id>/students/<int:student_id>/unlink', methods=['POST'])
@admin_required
def university_student_unlink(univ_id, student_id):
    univ = db.get_or_404(University, univ_id)
    student = User.query.filter_by(id=student_id, university_id=univ_id, role=ROLE_STUDENT).first_or_404()
    student.university_id = None
    student.university_name = None
    student.university_department_id = None
    student.university_class = None
    _audit('university.student_unlink', f'{student.email} ← {univ.name}')
    db.session.commit()
    flash('Student unlinked from university.', 'success')
    return _admin_university_student_redirect(univ_id)


@admin_bp.route('/universities/<int:univ_id>/students/<int:student_id>/delete', methods=['POST'])
@admin_required
def university_student_delete(univ_id, student_id):
    univ = db.get_or_404(University, univ_id)
    student = User.query.filter_by(id=student_id, university_id=univ_id, role=ROLE_STUDENT).first_or_404()
    if student.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return _admin_university_student_redirect(univ_id)

    name = student.full_name
    _delete_user_with_related_data(student)
    _audit('university.student_delete', f'{name} @ {univ.name}')
    db.session.commit()
    flash(f'Student "{name}" permanently deleted.', 'success')
    return _admin_university_student_redirect(univ_id)
